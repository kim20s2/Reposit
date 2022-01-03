import pandas as pd
import openpyxl
from openpyxl.styles.fonts import Font
from openpyxl import Workbook
import subprocess
import math
import time
import os
import re
import chardet
from enum import Enum
import datetime
import csv

start = time.time()

user_id_pw = 0
implementation_criteria = 0
minor_criteria = 0
test_result = 0
LukID_temp = 0
DecomposesToID = 0
SysID_length = 0
SwID_length = 0
cr_short_description = 0
cr_delivery_milestone = 0
verification_status = 'not finished'
TC_Review_Status = 0

def main():

    global row_cr
    global row_tc
    global user_id_pw
    global implementation_criteria
    global minor_criteria
    global LukID_temp
    global SwID_length
    global SysID_length
    global cr_delivery_milestone
    global TC_Review_Status

    implementation_criteria = 12
    row_temp_first = 2
    row_temp = 2
    minor_criteria = 'Patch#2'

    user_id_pw = [' ', ' ']
    # Get test session IDs

    user_name_txt = "user.txt"
    DocID_Info_xls = "DocID_Info.xls"
    SysID_Info_xls = "SysID_Info.xls"
    SwID_Info_xls = "SwID_Info.xls"
    SysSwTS_Info_xls = "SysSwTSID_Info.xls"                                 # 고정

    read_xlsx = "read_v104.7.xls"                                           # 변경
    test_session_txt = "testsession_v104.7.txt"                             # 변경
    TestResult_csv = "TestResult_Info_v104.7.csv"                           # 변경
    ########################## testsession.txt 파일 읽기 ##########################
    try:
        with open(test_session_txt, 'rt') as in_file:
            for line in in_file:
                test_session_id = line  # test_session_txt 파일에 1줄마다 읽어드림
                test_session_id = test_session_id.split(',')  # ,로 문자열 리스트화에서 구별
            len_test_session_id = len(test_session_id)
    except:
        print("Please make a testsession.txt file with test session id. e.g. 1495334, 1495339, 1495343, 1495344")  # test_seesion_txt 파일에 아무것도 데이터가 없을때 또는 ,로 구분안갈때
        subprocess.check_output("pause", shell=True)
        return

    ########################## user.txt 파일 읽기 ##########################
    try:
        with open(user_name_txt, 'rt') as in_file:
            for line in in_file:
                user_id_pw = line                               # user_name_txt 파일에 데이터 1줄마다 읽음
                user_id_pw = user_id_pw.split(',')              # 데이터의 id와 pw ,로 구분
                user_id_pw[0] = user_id_pw[0].strip()           # user_id_pw[0]에 id 데이터 저장
                user_id_pw[1] = user_id_pw[1].strip()           # user_id_pw[1]에 pw 데이터 저장
    except:
        print("Please make a user.txt file with ID, PW. e.g. ID, PW")
        subprocess.check_output("pause", shell=True)
        return

    row_cr = row_tc = 2   # 2행부터 시작
    ########################## read.xlsx 파일 읽기 ##########################

    worksheet_number = 0                                # sheet number 지정
    worksheet_name = "Sheet" + str(worksheet_number)    # read.xlsx의 worksheet 이름

    data1 = pd.read_excel(read_xlsx, sheet_name=worksheet_name)                     # read 파일 읽기 1~6
    df_read = pd.DataFrame(data1)

    data2 = pd.read_excel(DocID_Info_xls, sheet_name=worksheet_name)
    df_DocID = pd.DataFrame(data2)

    data3 = pd.read_excel(SysID_Info_xls, sheet_name=worksheet_name)
    df_SysID = pd.DataFrame(data3)

    data4 = pd.read_excel(SwID_Info_xls, sheet_name=worksheet_name)
    df_SwID = pd.DataFrame(data4)

    data5 = pd.read_excel(SysSwTS_Info_xls, sheet_name=worksheet_name)
    df_SysSwTS = pd.DataFrame(data5)

    data6 = pd.read_csv(TestResult_csv)
    df_TestResult = pd.DataFrame(data6)

    wb_write = Workbook()  # write.xlsx 파일 임시 파일 생성

    ws = wb_write.active
    ws.title = worksheet_name

    excel_header = ['CR Text', 'CR ID', 'ASIL Level', 'Acceptance LG against LuK requirement', 'Planned implementation milestone/date',
                    'Implementation state', 'If test: Test ID', 'Verification method',
                    'Test case review status', 'Test result' + str(test_session_id), 'Verification status', 'SysRS ID', 'SwRS ID', 'All TC Pass Status',
                    'If NOK or test not possible, add rational and write action planned']
    if (len(df_read) < 10000):
        ws.append(excel_header)

    else:                                                            # write.xlsx에 입력해야할 행이 10000개 이상일시 sheet를 한개 더 추가
        worksheet_number += 1
        worksheet_name = "Sheet" + str(worksheet_number)
        wb_write.create_sheet(worksheet_name)
        ws_write = wb_write[worksheet_name]
        ws.append(excel_header)

    #### 데이터 입력 시작 ####
    df_LuKID = df_read['A15 LuK ID']
    p = re.compile('M([\d]*)')                                      # 정규표현식을 미리 컴파일 해두는것
    for j in range(0, len(df_read)):                                # read파일에 있는 A15 LuK ID의 갯수만큼 for문
        All_TC_Pass_data = []
        DocID_Info_line = data2["A15 LuK ID"].str.contains(df_LuKID[j])
        Save_DocID_Info_line = data2[DocID_Info_line]
        if Save_DocID_Info_line.empty:
            data_row = ['WRONG CR ID', 'WRONG CR ID']
            ws.append(data_row)
            row_temp += 1
            continue
        data_row = Save_DocID_Info_line.loc[:, ['Text', "A15 LuK ID", "A05 Safety Integrity", "A25 Status Commitment Supplier - MCA LG", "A27 Delivery Date"]]
        data_row = data_row.values.tolist()[0]                         # Text, A15 LuK ID, A05 Safety Integrity, A25 Status Commitment Supplier - MCA LG, A27 Delivery Date 추출 완료//

        if str(data_row[2]) == 'nan':
            data_row[2] = 'None'

        if str(data_row[4]) == 'nan':                                          # data_row[4] 는 A27 Delivery Date에 대한 정보 M10, M11, n/a 같은 // str(data_row[4]) == 'nan' 로 바꿔야 하나?
            data_row[4] = 'in-review'
        elif data_row[4] =='all milestones':
            data_row[4] = 'all milestones'
        else:
            if data_row[4] != 'n/a':
                m = p.search(data_row[4])
                cr_delivery_milestone = m.group(1)
            else:
                cr_delivery_milestone = 1

        if (int(cr_delivery_milestone) != 0) & (int(cr_delivery_milestone) <= implementation_criteria):                      #implementation_criteria = 12 , cr_delivery_milesone이 0이 아니고 숫자가 12보다 작은 경우
            if minor_criteria in str(data_row[4]):                                                                           #cr_delivery = M10, M11, n/a같은 / minor_criteria = 'Patch#2'
                cr_delivery_milestone = 'not implemented'
            else:
                cr_delivery_milestone = 'implemented'                                                                        #거의 다 이쪽 실행
        else:
            cr_delivery_milestone = 'not implemented'

        data_row.append(cr_delivery_milestone)                           # cr_delivery_milestone 추출 완료//
        cr_delivery_milestone = 0

        Decom_temp = Save_DocID_Info_line.loc[:, ["Decomposes To"]]      # DecomposesTo ID 임시 저장소
        Decom_temp = Decom_temp.values.tolist()[0]
        DecomID_data_row = Decom_temp[0]

        Short_temp = Save_DocID_Info_line.loc[:, ["Short Description"]]
        Short_temp = Short_temp.values.tolist()[0]
        Short_temp = Short_temp[0]

        if str(DecomID_data_row) != 'nan':
            DecomID_data_row = DecomID_data_row.split(',')
            DecomID_length = len(DecomID_data_row)
            for i in range(0, DecomID_length):
                DecomposesToID = DecomID_data_row[i].replace("?", "").lstrip().rstrip()
                SysID_Info_line = df_SysID.loc[df_SysID['ID'] == int(DecomposesToID)]
                if SysID_Info_line.empty:
                    #ws.append(data_row)
                    #row_temp += 1
                    continue
                SysID_data_row = SysID_Info_line.loc[:, ["ENG ID", "Validated By", "Satisfied By"]]
                SysID_data_row = SysID_data_row.values.tolist()[0]

                SysRSID = SysID_data_row[0]
                SysTsID_data_row = SysID_data_row[1]
                SwID_row = SysID_data_row[2]

                if str(SysTsID_data_row) != 'nan':
                    SysTsID_data_row = SysTsID_data_row.split(',')
                    SysTsID_length = len(SysTsID_data_row)
                    for i in range(0, SysTsID_length):
                        SysTsID = SysTsID_data_row[i].replace("?", "").lstrip().rstrip()
                        SysTSID_Info_line = df_SysSwTS.loc[df_SysSwTS['ID'] == int(SysTsID)]
                        SysTSID_data_row = SysTSID_Info_line.loc[:, ["ENG ID", "Test Method"]]
                        if SysTSID_data_row.empty:
                            SysTSID_data_row = ['n/a', 'n/a']
                            TC_Review_Status = 'n/a'
                        else:
                            SysTSID_data_row = SysTSID_data_row.values.tolist()[0]
                            TC_Review_Status = 'reviewed'

                        SysTSID_temp = SysTSID_Info_line.loc[:, ["ID"]]
                        if SysTSID_temp.empty:
                            TC_Result = 'NT'
                        else:
                            SysTSID_temp = SysTSID_temp.values.tolist()[0][0]
                            TC_Result_Info_line = df_TestResult.loc[df_TestResult["TC ID"] == int(SysTSID_temp)]
                            TC_Result_data_row = TC_Result_Info_line.loc[:, ["Session ID", "Test Result"]]
                            if TC_Result_data_row.empty:
                                TC_Result = 'NT'
                            else:
                                TC_Result_data_row = TC_Result_data_row.values.tolist()[0]
                                TC_Result = TC_Result_data_row[1] + "(session : " + str(TC_Result_data_row[0]) + ")"

                        row_data = data_row + SysTSID_data_row
                        row_data.append(TC_Review_Status)
                        row_data.append(TC_Result)

                        if str('Passed') in str(TC_Result):
                            All_TC_Pass_data.append('P')
                        else:
                            All_TC_Pass_data.append('F')

                        if str('Passed') in str(TC_Result):
                            verification_status = 'finished'
                        elif str('Failed') in str(TC_Result):
                            verification_status = 'not finished'
                        else:
                            if Short_temp == 'finished':
                                verification_status = 'finished'
                            else:
                                verification_status = 'not finished'
                        row_data.append(verification_status)
                        row_data.append(SysRSID)
                        ws.append(row_data)
                        row_temp += 1

                if str(SwID_row) != 'nan':
                    SwID_row = SwID_row.split(',')
                    SwID_row = [i.replace('?', '').strip() for i in SwID_row]
                    SwID_length = len(SwID_row)
                    for i in range(0, SwID_length):
                        SwID = SwID_row[i]
                        SwID_Info_line = df_SwID.loc[df_SwID['ID'] == int(SwID)]
                        SwID_data_row = SwID_Info_line.loc[:, ["ENG ID", "Validated By"]]
                        if SwID_Info_line.empty:
                            #row_data = data_row + [' ', ' ', ' ', ' ', ' ', SysRSID]
                            #ws.append(row_data)
                            #row_temp += 1
                            continue
                        else:
                            SwID_data_row = SwID_data_row.values.tolist()[0]

                        SwRSID = SwID_data_row[0]
                        SwTsID_data_row = SwID_data_row[1]

                        if str(SwTsID_data_row) != 'nan':
                            SwTsID_ValidatedBy = SwTsID_data_row.split(",")
                            SwTsID_length = len(SwTsID_ValidatedBy)
                            for i in range(0, SwTsID_length):
                                SwTsID = SwTsID_ValidatedBy[i].replace("?", "").lstrip().rstrip()
                                SwTSID_Info_line = df_SysSwTS.loc[df_SysSwTS['ID'] == int(SwTsID)]
                                SwTSID_data_row = SwTSID_Info_line.loc[:, ["ENG ID", "Test Method"]]
                                if SwTSID_data_row.empty:
                                    SwTSID_data_row = ['n/a', 'n/a']
                                    TC_Review_Status = 'n/a'
                                else:
                                    SwTSID_data_row = SwTSID_data_row.values.tolist()[0]
                                    TC_Review_Status = 'reviewed'

                                SwTSID_temp = SwTSID_Info_line.loc[:, ["ID"]]
                                if SwTSID_temp.empty:
                                    TC_Result = 'NT'
                                else:
                                    SwTSID_temp = SwTSID_temp.values.tolist()[0][0]
                                    TC_Result_Info_line = df_TestResult.loc[df_TestResult["TC ID"] == int(SwTSID_temp)]
                                    TC_Result_data_row = TC_Result_Info_line.loc[:, ["Session ID", "Test Result"]]
                                    if TC_Result_data_row.empty:
                                        TC_Result = 'NT'
                                    else:
                                        TC_Result_data_row = TC_Result_data_row.values.tolist()[0]
                                        TC_Result = TC_Result_data_row[1] + "(session : " + str(TC_Result_data_row[0]) + ")"

                                row_data = data_row + SwTSID_data_row
                                row_data.append(TC_Review_Status)
                                row_data.append(TC_Result)

                                if str('Passed') in str(TC_Result):
                                    All_TC_Pass_data.append('P')
                                else:
                                    All_TC_Pass_data.append('F')

                                if str('Passed') in str(TC_Result):
                                    verification_status = 'finished'
                                elif str('Failed') in str(TC_Result):
                                    verification_status = 'not finished'
                                else:
                                    if Short_temp == 'finished':
                                        verification_status = 'finished'
                                    else:
                                        verification_status = 'not finished'

                                row_data.append(verification_status)
                                row_data.append(' ')
                                row_data.append(SwRSID)
                                ws.append(row_data)
                                row_temp += 1

        if 'F' in All_TC_Pass_data:
            All_TC_Pass = "All TC Not Pass"
        else:
            All_TC_Pass = "All TC Pass"

        if j == 0:
            row_temp = row_temp_first

        ws.cell(row_temp, 14).value = All_TC_Pass



    #### 데이터 입력 종료 ####
    wb_write.save("TraceMatrix_v104.7.xlsx")
    wb_write.close()

main()

end = time.time()
print(f"Write파일이 만드는데 까지 걸리는 시간 : {end - start:.5f} sec")