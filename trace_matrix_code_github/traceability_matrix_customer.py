import openpyxl
from openpyxl import Workbook
# coding=utf-8
import os
import re
from enum import Enum
import datetime
import time
import subprocess
import chardet
import getTestResult                                                     #getTestResult.py import

from openpyxl.utils.dataframe import dataframe_to_rows

#import logging                                                          #시간에 영향을 줌
#import logging.handlers                                                 #시간에 영향을 줌

row_cr = 0                                                               #######전역 변수######
row_tc = 0
user_id_pw = 0
implementation_criteria = 0
minor_criteria = 0
skip_string = 0
id_decom = 0
engid = 0
luk_name = 0
id_validate = 0
id_satis = 0
gurl = 0
test_result = 0

now = datetime.datetime.now()
today = now.strftime('%Y-%m-%d %H:%M')
nowtime = now.strftime('%m%d_%H%M')
outputFileName_MCA_Full='liveDocTraceInfo_MCA_Full_'+nowtime+'.xls'

wb_data = Workbook()
wb_data.save("data.xlsx")
ws_data = wb_data.active

class idx_id(Enum):                                                        #클래스 변수 생성
    IDX_CR_TEXT = 1
    IDX_CR_ID = 2
    IDX_ASIL = 3
    IDX_ACCEPTANCE_LG = 4
    IDX_IMPLEMENT_STATUS = 5
    IDX_IMPLEMENT_MILESTONE = 6
    IDX_VC_METHOD = 7
    IDX_TC = 8
    IDX_TC_REVIEW_STATUS = 9
    IDX_VC_STATUS = 10
    IDX_RESULT = 11
    IDX_ACTION_PLAN = 12
    IDX_SYSRS_ID = 13
    IDX_SWRS_ID = 14

class _RegExLib_mileston:
    """Set up regular expressions"""
    # use https://regexper.com to visualise these if required
    _reg_delivery_milestone = re.compile('A27 Delivery Date: M([\d]*)')                                     #정규표현식을 미리 컴파일 해두는것

    def __init__(self, line):
        # check whether line has a positive match with all of the regular expressions
        self.delivery_milestone = self._reg_delivery_milestone.match(line)

class _RegExLib:
    """Set up regular expressions"""
    # use https://regexper.com to visualise these if required
    _reg_engid = re.compile('ENG ID: (.*)')                                             # _reg_engid에 'ENG ID: ~' 라는 문자열 패턴을 모두 저장 & 컴파일
    _reg_decomposes = re.compile('Decomposes To: (.*)')
    _reg_luk = re.compile('A15 LuK ID: (.*)')
    _reg_validates = re.compile('Validated By: (.*)')
    _reg_satisfies = re.compile('Satisfied By: (.*)')
    _reg_url = re.compile('URL: (.*)')
    _reg_text = re.compile('Text: (.*)')
    _reg_asil = re.compile('A05 Safety Integrity: (.*)')
    _reg_acceptance_lg = re.compile('A25 Status Commitment Supplier - MCA LG: (.*)')
    _reg_acceptance_luk = re.compile('A25 Status Commitment LuK \(as customer\): (.*)')
    _reg_delivery = re.compile('A27 Delivery Date: (.*)')
    _reg_verification_method = re.compile('Test Method: (.*)')
    _reg_short_description = re.compile('Short Description: (.*)')

    def __init__(self, line):
        # check whether line has a positive match with all of the regular expressions
        self.engid = self._reg_engid.match(line)
        self.decomposes = self._reg_decomposes.match(line)
        self.luk = self._reg_luk.match(line)
        self.validates = self._reg_validates.match(line)
        self.satisfies = self._reg_satisfies.match(line)
        self.url = self._reg_url.match(line)
        self.text = self._reg_text.match(line)
        self.asil = self._reg_asil.match(line)
        self.acceptance_lg = self._reg_acceptance_lg.match(line)
        self.acceptance_luk = self._reg_acceptance_luk.match(line)
        self.delivery = self._reg_delivery.match(line)
        self.verification_method = self._reg_verification_method.match(line)
        self.short_description = self._reg_short_description.match(line)

def get_luk_name(id) -> object:
    # view_issue_command = 'im viewissue --user=' + str(user_id_pw[0]) + ' --password='+str(user_id_pw[1]) + ' '        #user.txt파일에서 얻은 user id와 pw입력 :im viewissue --user=kim30s2 --password=kim30s2

    view_issue_command = 'im viewissue' + ' '
    im_command_txt = view_issue_command + id                                                                          #im viewissue --user=kim30s2 --password=kim30s2 1228293
    result = subprocess.check_output(im_command_txt, shell=True)
    result = result.splitlines()


    cr_text = 0
    need_to_skip = 0
    name_decom_split = 0
    cr_asil = 0
    cr_acceptance_lg = 0
    cr_acceptance_luk = 0
    cr_delivery = 0
    cr_verification_method = 0
    cr_delivery_milesone = 0
    cr_short_description = 0
    cr_url = 0

    for line in result:  # Store each line in a string variable "line"
        # parse id decomposed to, satisfied by, validated by
        try:
            encode_type = chardet.detect(line)
            if encode_type['encoding'] is not None:
                line = line.decode(encode_type['encoding'])
            else:
                line = line.decode('EUC-KR')                                                    #EUC-KR형식으로 byte코드를 문자열로 반환
            reg_match = _RegExLib(line)                                                         #reg_match에 클라스 유형으로 _RegExLib(line)을 저장
        except:
            print("problem is occured. id", id)
            print(line)

        if reg_match.luk:
            # write to excel to CR ID column
            name_decom = reg_match.luk.group(1)                                                 #reg_match.luk <-- self.luk = self._reg_luk.match(line) <-- _reg_luk = re.compile('A15 LuK ID: (.*)')
            if name_decom == '':
                name_decom_split = 0
            else:
                name_decom_split = name_decom.split(',')                                        #문자열을 ,로 나누어서 리스트화

        if reg_match.url:
            cr_url = reg_match.url.group(1)                                                     #reg_match.url에서 첫번째 그룹에 해당되는 문자열을 cr_url로 반환

        if reg_match.text:
            cr_text = reg_match.text.group(1)                                                   #reg_match.text에서 첫번째 그룹에 해당되는 문자열을 cr_text로 반환
            try:
                cr_text = cr_text.decode('iso-8859-1').encode("utf-8").strip()                  #iso-8859-1 형식으로 cr_text를 문자열로 반환후 utf-8형식으로 다시 그 문자열을 숫자로 반환
            except:
                cr_text = cr_text.strip()                                                       #문자열의 젤 왼쪽과 오른쪽의 공백 제거

        if reg_match.asil:
            cr_asil = reg_match.asil.group(1)                                                   #reg_match.asil 문자열의 첫번째 그룹을 반환
            if cr_asil == '':
                cr_asil = 'None'

        if reg_match.acceptance_lg:
            cr_acceptanfce_lg = reg_match.acceptance_lg.group(1)
            if cr_acceptance_lg == '':
                cr_acceptance_lg = 'to write acceptance'
            elif cr_acceptance_lg == 'rejected':
                need_to_skip = 1

        if reg_match.acceptance_luk:
            cr_acceptance_luk = reg_match.acceptance_luk.group(1)
            if cr_acceptance_luk == 'canceled':
                need_to_skip = 1

        if reg_match.delivery:
            cr_delivery = reg_match.delivery.group(1)
            if cr_delivery == '':
                cr_delivery = 'in-review'
            elif cr_delivery =='all milestones':
                cr_delivery = 'all milestones'
            else:
                if cr_delivery != 'n/a':
                    reg_match_milestone = _RegExLib_mileston(line)
                    cr_delivery_milesone = reg_match_milestone.delivery_milestone.group(1)
                else:
                    cr_delivery_milesone = 1

        if reg_match.verification_method:                                               # 스킵되는듯
            cr_verification_method = reg_match.verification_method.group(1)
            if cr_verification_method == '':
                cr_verification_method = 'to write verification method'

        if reg_match.short_description:
            cr_short_description = reg_match.short_description.group(1)
            if cr_short_description == 'finished':
                cr_short_description = 1

    return name_decom_split, cr_url, cr_text, cr_asil, cr_acceptance_lg, cr_delivery, cr_verification_method, cr_delivery_milesone, need_to_skip, cr_short_description

def get_decom_ids(id):                              #id means session id
    #remove white space and ,
    id = id.strip()                                                         # 문자열 공백 제거 ' Water ' -> 'Water'
    id = id.replace('?','')

    # view_issue_command = 'im viewissue --user=' + str(user_id_pw[0]) + ' --password='+str(user_id_pw[1]) + ' '
    view_issue_command = 'im viewissue' + ' '
    im_command_txt = view_issue_command + id
    result = subprocess.check_output(im_command_txt, shell=True)
    result = result.splitlines()                                            # 줄바꿈\n이 있는 곳을 기준으로 리스트로 나누어 반환 'a\nb' -> ['a', 'b']

    ldecom_array = lengid = lluk_name = lid_validate = lid_satis = lurl = cr_verification_method = 0
    ldecom_array = []
    lid_satis = []
    get_decom_ids_return = []

    for line in result:  # Store each line in a string variable "line"
        # parse id decomposed to, satisfied by, validated by
        try:
            encode_type = chardet.detect(line)
            if encode_type['encoding'] is not None:
                line = line.decode(encode_type['encoding'])
            else:
                line = line.decode('EUC-KR')
            reg_match = _RegExLib(line)

        except:
            print("problem is occured. id", id)
            print(line)

        if reg_match.decomposes:
            ldecom_array = reg_match.decomposes.group(1)
            ldecom_array = ldecom_array.split(',')

        if reg_match.engid:
            lengid = reg_match.engid.group(1)

        if reg_match.luk:
            lluk_name = reg_match.luk.group(1)

        if reg_match.validates:
            lid_validate = reg_match.validates.group(1)
            lid_validate = lid_validate.split(',')

        if reg_match.satisfies:
            lid_satis = reg_match.satisfies.group(1)
            lid_satis = lid_satis.split(',')

        if reg_match.url:
            lurl = reg_match.url.group(1)

        if reg_match.verification_method:
            cr_verification_method = reg_match.verification_method.group(1)
            if cr_verification_method == '':
                cr_verification_method = 'to write verification method'

    return ldecom_array, lengid, lluk_name, lid_validate, lid_satis, lurl, cr_verification_method


def main():
    global row_cr
    global row_tc
    global user_id_pw
    global implementation_criteria
    global minor_criteria
    global skip_string

    implementation_criteria = 12
    minor_criteria = 'Patch#2'

    user_id_pw = [' ',' ']
    # Get test session IDs
    test_session_txt = "testsession.txt"
    user_name_txt = "user.txt"

    try:
        with open(test_session_txt, 'rt') as in_file:
            for line in in_file:
                test_session_id = line                                          #test_session_txt 파일에 1줄마다 읽어드림
                test_session_id = test_session_id.split(',')                    #,로 문자열 리스트화에서 구별
    except:
        print("Please make a testsession.txt file with test session id. e.g. 1495334, 1495339, 1495343, 1495344")               #test_seesion_txt 파일에 아무것도 데이터가 없을때 또는 ,로 구분안갈때
        subprocess.check_output("pause", shell=True)
        return

    try:
        with open(user_name_txt, 'rt') as in_file:
            for line in in_file:
                user_id_pw = line                                               #user_name_txt 파일에 데이터 1줄마다 읽음
                user_id_pw = user_id_pw.split(',')                              #데이터의 id와 pw ,로 구분
                user_id_pw[0] = user_id_pw[0].strip()                           #user_id_pw[0]에 id 데이터 저장
                user_id_pw[1] = user_id_pw[1].strip()                           #user_id_pw[1]에 pw 데이터 저장
    except:
        print("Please make a user.txt file with ID, PW. e.g. ID, PW")
        subprocess.check_output("pause", shell=True)
        return

    row_cr = row_tc = 2
    ########################## 엑셀파일 ############################
    try:
        wb_read = openpyxl.load_workbook('read.xlsx')
    except:
        print("Please make a read.xlsx file with CR ID there are A column")
        subprocess.check_output("pause", shell=True)
        return

    worksheet_number = 1
    worksheet_name = "Sheet" + str(worksheet_number)                                            #read.xlsx의 worksheet 네이밍과 번호
    wb_write = openpyxl.Workbook()


    try:
        subprocess.check_output("rm write1.xlsx", shell=True)
        wb_write.create_sheet(worksheet_name)
        ws_write = wb_write["Sheet"]
        wb_write.remove(ws_write)
        wb_write.save('write1.xlsx')
    except:
        wb_write.create_sheet(worksheet_name)
        ws_write = wb_write["Sheet"]
        wb_write.remove(ws_write)
        wb_write.save('write1.xlsx')


    # 현재 Active Sheet 얻기
    # ws = wb_read.active

    ws = wb_read["Sheet0"]                                                      #ws은 read.xlsx파일의 Sheet0 데이터
    ws_write = wb_write[worksheet_name]


    # Read cr id
    row_cr = 2                                                                                      #row_cr은 무슨뜻??
    ## server에 login

    connect_server_command = 'si connect --user=' + str(user_id_pw[0]) + ' --password=' + str(user_id_pw[1])
    subprocess.call(connect_server_command)

    for r in ws.rows:                                                   #read.xlsx파일 읽어보기
        if r[0].row == 1:                                               #ws파일의 첫번째 행일때         // r[0]는 ws데이터의 첫번째 행의 첫번째 열의 데이터이다

            ws_write.append(['CR Text', 'CR ID', 'ASIL Level', 'acceptance LG against LuK requirement', 'implementation state', 'planned implementation milestone/date', 'verification method', 'if test: Test ID', 'test case review status', 'verification status', 'test result' + str(test_session_id), 'If NOK or test not possible, add rational and write action planned'])
            continue
        # row_index = r[0].row
        if row_cr > 10000:                                                  #write.xlsx에 입력해야할 행이 10000개 이상일시 sheet를 한개 더 추가
            worksheet_number += 1
            worksheet_name = "Sheet" + str(worksheet_number)
            wb_write.create_sheet(worksheet_name)
            ws_write = wb_write[worksheet_name]
            ws_write.cell(row=1, column=idx_id.IDX_CR_TEXT.value).value = 'CR Text'
            ws_write.cell(row=1, column=idx_id.IDX_CR_ID.value).value = 'CR ID'
            ws_write.cell(row=1, column=idx_id.IDX_ASIL.value).value = 'ASIL Level'
            ws_write.cell(row=1, column=idx_id.IDX_ACCEPTANCE_LG.value).value = 'acceptance LG against LuK requirement'
            ws_write.cell(row=1, column=idx_id.IDX_IMPLEMENT_STATUS.value).value = 'implementation state'
            ws_write.cell(row=1, column=idx_id.IDX_IMPLEMENT_MILESTONE.value).value = 'planned implementation milestone/date'
            ws_write.cell(row=1, column=idx_id.IDX_VC_METHOD.value).value = 'verification method'
            ws_write.cell(row=1, column=idx_id.IDX_TC.value).value = 'if test: Test ID'
            ws_write.cell(row=1, column=idx_id.IDX_TC_REVIEW_STATUS.value).value = 'test case review status'
            ws_write.cell(row=1, column=idx_id.IDX_VC_STATUS.value).value = 'verification status'
            ws_write.cell(row=1, column=idx_id.IDX_RESULT.value).value = 'test result' + str(test_session_id)
            ws_write.cell(row=1, column=idx_id.IDX_ACTION_PLAN.value).value = 'If NOK or test not possible, add rational and write action planned'
            row_cr = row_tc = 2

        cr_id = r[0].value                                  #read파일에서 첫번째 열인 ID에 있는 숫자 첫번째값
        cr_id = str(cr_id)
        # if row_cr != 2 :
        #     row_cr = row_tc
        cr_luk_name, cr_url, cr_text, cr_asil, cr_acceptance_lg, cr_delivery, cr_verification_method, cr_delivery_milesone, cr_need_to_skip, cr_short_description = get_luk_name(cr_id)

        if cr_need_to_skip == 1:
            continue
        if (int(cr_delivery_milesone) != 0) & (int(cr_delivery_milesone) <= implementation_criteria):                           #implementation_criteria = 12 , cr_delivery_milesone이 0이 아니고 숫자가 12보다 작은 경우
            if minor_criteria in str(cr_delivery):                                                                              #cr_delivery = M10, M11, n/a같은 / minor_criteria = 'Patch#2'
                cr_delivery_milesone = 'not implemented'
            else:
                cr_delivery_milesone = 'implemented'                                                                        #거의 다 이쪽 실행
        else:
            cr_delivery_milesone = 'not implemented'


        ws_write.cell(row=row_cr, column=idx_id.IDX_CR_TEXT.value).value = str(cr_text)                                 #row_cr의 열에 cr_text
        ws_write.cell(row=row_cr, column=idx_id.IDX_CR_ID.value).value = str(cr_luk_name)
        ws_write.cell(row=row_cr, column=idx_id.IDX_CR_ID.value).hyperlink = str(cr_url)
        ws_write.cell(row=row_cr, column=idx_id.IDX_ASIL.value).value = str(cr_asil)
        ws_write.cell(row=row_cr, column=idx_id.IDX_ACCEPTANCE_LG.value).value = str(cr_acceptance_lg)
        ws_write.cell(row=row_cr, column=idx_id.IDX_IMPLEMENT_STATUS.value).value = str(cr_delivery_milesone)
        ws_write.cell(row=row_cr, column=idx_id.IDX_IMPLEMENT_MILESTONE.value).value = str(cr_delivery)
        ws_write.cell(row=row_cr, column=idx_id.IDX_VC_METHOD.value).value = str(cr_verification_method)

        print(str(datetime.datetime.now()) + str(cr_luk_name) + 'row number = ' + str(r[0].row ))                       #cmd 창에서 뜨는 처리 상황

        row_tc = row_cr
        row_cr += 1
        start_group_id = row_cr
        global id_decom                             #함수(main)안에서도 전역변수들을 수정하기 위해서 global로 선언
        global engid
        global luk_name
        global id_validate
        global id_satis
        global gurl
        global test_result
        id_decom = engid = luk_name = id_validate = id_satis = 0
        skip_string = 'TSR_'
        # get ids of CR

        id_decom, engid, luk_name, id_validate, id_satis, gurl, cr_verification_method = get_decom_ids(cr_id)

        # check each ids
        if id_decom != ['']:
            # there are sysrs ids
            for each_sysrs_id in id_decom:
                # get ids of Sysrs
                decompose_id_sysrs, engid_sysrs, luk_name_sysrs, validate_id_sysrs, satisfies_id_sysrs, gurl_sysrs, cr_verification_method = get_decom_ids(each_sysrs_id)
                if str(engid_sysrs) == '':
                    engid_sysrs = 'need to type ENG ID'
                # pseudo code
                # engid 만약에 TSR로 시작하면 Skip하자(continue)
                if skip_string in str(engid_sysrs):
                    continue
                ws_write.cell(row=row_tc, column=idx_id.IDX_SYSRS_ID.value).value = str(engid_sysrs)
                ws_write.cell(row=row_tc, column=idx_id.IDX_SYSRS_ID.value).hyperlink = str(gurl_sysrs)
                #row_sysrs += 1
                if validate_id_sysrs != ['']:   ######################################## validate ##############################################
                    for each_valid_id in validate_id_sysrs:
                        id_systs_array, engid_systs, luk_name_systs, validate_id_systs, satisfies_id_systs, gurl_systs, cr_verification_method = get_decom_ids(each_valid_id)
                        if str(engid_systs) == '':
                            engid_systs = 'need to type ENG ID'

                        ws_write.cell(row=row_tc, column=idx_id.IDX_TC.value).value = str(engid_systs)
                        ws_write.cell(row=row_tc, column=idx_id.IDX_TC.value).hyperlink = str(gurl_systs)
                        ws_write.cell(row=row_tc, column=idx_id.IDX_TC_REVIEW_STATUS.value).value = 'reviewed'
                        ws_write.cell(row=row_tc, column=idx_id.IDX_VC_METHOD.value).value = str(cr_verification_method)

                        test_result = getTestResult.GetTestResultSessions(test_session_id, each_valid_id)

                        ws_write.cell(row=row_tc, column=idx_id.IDX_RESULT.value).value = str(test_result)

                        verification_status = 'not finished'
                        if str('Passed') in str(test_result):
                            verification_status = 'finished'
                        elif str('Failed') in str(test_result):
                            verification_status = 'not finished'
                        else:
                            if cr_short_description == 1:
                                ws_write.cell(row=row_tc, column=idx_id.IDX_VC_STATUS.value).value = 'finished'
                            else:
                                verification_status = 'not finished'

                        ws_write.cell(row=row_tc, column=idx_id.IDX_VC_STATUS.value).value = str(verification_status)
                        row_tc += 1

                # there are swrs
                if satisfies_id_sysrs != ['']:  ######################################## satisfies ##############################################
                    for each_swrs_id in satisfies_id_sysrs:
                        id_swrs_array, engid_swrs, luk_name_swrs, validate_id_swrs, satisfies_id_swrs, gurl_swrs, cr_verification_method = get_decom_ids(each_swrs_id)
                        if str(engid_swrs) == '':
                            engid_swrs = 'need to type ENG ID'

                        ws_write.cell(row=row_tc, column=idx_id.IDX_SWRS_ID.value).value = str(engid_swrs)
                        ws_write.cell(row=row_tc, column=idx_id.IDX_SWRS_ID.value).hyperlink = str(gurl_swrs)
                        #row_swrs += 1
                        if validate_id_swrs != ['']:
                            for each_valid_id in validate_id_swrs:
                                id_swts_array, engid_swts, luk_name_swts, validate_id_swts, satisfies_id_swts, gurl_swts, cr_verification_method = get_decom_ids(each_valid_id)

                                if str(engid_swts) == '':
                                    engid_swts = 'need to type ENG ID'

                                ws_write.cell(row=row_tc, column=idx_id.IDX_TC.value).value = str(engid_swts)
                                ws_write.cell(row=row_tc, column=idx_id.IDX_TC.value).hyperlink = str(gurl_swts)
                                ws_write.cell(row=row_tc, column=idx_id.IDX_TC_REVIEW_STATUS.value).value = 'reviewed'
                                ws_write.cell(row=row_tc, column=idx_id.IDX_VC_METHOD.value).value = str(cr_verification_method)
                                test_result = getTestResult.GetTestResultSessions(test_session_id, each_valid_id)

                                ws_write.cell(row=row_tc, column=idx_id.IDX_RESULT.value).value = str(test_result)
                                verification_status = 'not finished'
                                if str('Passed') in str(test_result):
                                    verification_status = 'finished'
                                elif str('Failed') in str(test_result):
                                    verification_status = 'not finished'
                                else:
                                    if cr_short_description == 1:
                                        ws_write.cell(row=row_tc, column=idx_id.IDX_VC_STATUS.value).value = 'finished'
                                    else:
                                        verification_status = 'not finished'
                                ws_write.cell(row=row_tc, column=idx_id.IDX_VC_STATUS.value).value = str(verification_status)

                                row_tc += 1

        end_group_id = row_tc
        # row_cr와 row_tc 중 큰 값으로 row_cr로 정한다.
        if row_tc > row_cr:
            row_cr = row_tc
        if start_group_id != end_group_id:
            for i in range(start_group_id , end_group_id):
                ws_write.row_dimensions[i].outlineLevel = 1
                ws_write.row_dimensions[i].collapsed = True
        # 엑셀 파일 저장


    wb_write.save("write1.xlsx")                                                                     #write1.xlsx에 저장
    print("End!! Please type enter")
    subprocess.check_output("pause", shell=True)




main()                     #이 코드 꼭 필요 왜냐하면 main() 실행을 해줘야 하기 때문^