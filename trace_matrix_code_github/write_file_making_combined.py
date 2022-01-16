import pandas as pd
import openpyxl
from openpyxl.styles.fonts import Font
import win32com.client
from openpyxl import Workbook
import subprocess
import math
import time
import os
import re
import chardet
from enum import Enum
import datetime
import csv
import xlwt

start1 = time.time()

user_id_pw = 0
implementation_criteria = 0
minor_criteria = 0
test_result = 0
LukID_temp = 0
DecomposesToID = 0
SysID_length = 0
SwID_length = 0
cr_short_description = 0
cr_delivery_milestone = 0
verification_status = 'not finished'
TC_Review_Status = 0

DocID = '1226890'                                                       ############## 고정이나 나중에 변경될 수 있음
SysRS1ID = '1351133'
SysRS2ID = '1356188'
SysRS3ID = '1430240'
SwRS1ID = '1392803'
SwRS2ID = '1394006'
SwRS3ID = '1469578'
SysTCID = '1454824'
SwTCID = '1464826'
SysITSID = '1454310'

################################★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★################################
version = "104.7"                                                       ############### 여기만 변경
read_xlsx = "read_v" + version + ".xls"                                 ############### 파일명 format
test_session_txt = "testsession_v" + version + ".txt"
TestResult_csv = "TestResult_Info_v" + version + ".csv"
write_xlsx = "TraceMatrix_v" + version + ".xlsx"
################################★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★################################


def SysID_combined(input1, input2, input3):
    #엑셀 파일 이름
    excel_names = [input1, input2, input3]
    excels = [pd.ExcelFile(name) for name in excel_names]
    frames = [x.parse(x.sheet_names[0], header=None,index_col=None) for x in excels]
    frames[1:] = [df[1:] for df in frames[1:]]
    combined = pd.concat(frames)
    temp_list = (combined.iloc[0]).values.tolist()
    if 'ID' not in temp_list:
        header = pd.DataFrame([["Document ID",'ID',"A15 LuK ID",'Text',"A05 Safety Integrity","A25 Status Commitment Supplier - MCA LG","A27 Delivery Date","Decomposes To","Short Description"]])
        combined = pd.concat([header, combined], ignore_index=True)

    #파일저장
    combined.to_excel("SysID_Info.xlsx", header=False, index=False)

def SwID_combined(input1, input2, input3):
    #엑셀 파일 이름
    excel_names = [input1, input2, input3]
    excels = [pd.ExcelFile(name) for name in excel_names]
    frames = [x.parse(x.sheet_names[0], header=None,index_col=None) for x in excels]
    frames[1:] = [df[1:] for df in frames[1:]]
    combined = pd.concat(frames)
    temp_list = (combined.iloc[0]).values.tolist()
    if 'ID' not in temp_list:
        header = pd.DataFrame([["Document ID",'ID',"ENG ID","Validated By","Satisfied By"]])
        combined = pd.concat([header, combined], ignore_index=True)

    #파일저장
    combined.to_excel("SwID_Info.xlsx", header=False, index=False)

def SysSwTSID_combined(input1, input2, input3):
    #엑셀 파일 이름
    excel_names = [input1, input2, input3]
    excels = [pd.ExcelFile(name) for name in excel_names]
    frames = [x.parse(x.sheet_names[0], header=None,index_col=None) for x in excels]
    frames[1:] = [df[1:] for df in frames[1:]]
    combined = pd.concat(frames)
    temp_list = (combined.iloc[0]).values.tolist()
    if 'ID' not in temp_list:
        header = pd.DataFrame([["Document ID",'ID',"ENG ID","Test Method"]])
        combined = pd.concat([header, combined], ignore_index=True)

    #파일저장
    combined.to_excel("SysSwTSID_Info.xlsx", header=False, index=False)


def TestResult_Info():
    global test_session_txt
    global TestResult_csv

    try:
        with open(test_session_txt, 'rt') as in_file:
            for line in in_file:
                test_session_id = line  # test_session_txt 파일에 1줄마다 읽어드림
                test_session_id = test_session_id.split(',')  # ,로 문자열 리스트화에서 구별
            len_test_session_id = len(test_session_id)
    except:
        print("Please make a testsession.txt file with test session id. e.g. 1495334, 1495339, 1495343, 1495344")  # test_seesion_txt 파일에 아무것도 데이터가 없을때 또는 ,로 구분안갈때
        subprocess.check_output("pause", shell=True)
        return

    f = open(TestResult_csv, 'w', encoding='EUC-KR', newline='')
    wr = csv.writer(f)
    wr.writerow(["Session ID", "TC ID", "Test Result"])

    itemExportFields_TestResult = "sessionID,caseID,verdict"

    for j in range(0, len_test_session_id):
        result_test_cmd = 'tm results' + ' --sessionID=' + test_session_id[j].strip() +' --fields=' + itemExportFields_TestResult

        result = subprocess.check_output(result_test_cmd)
        result = result.splitlines()

        for line in result:  # Store each line in a string variable "line"
            # parse id decomposed to, satisfied by, validated by
            try:
                encode_type = chardet.detect(line)
                if encode_type['encoding'] is not None:
                    line = line.decode(encode_type['encoding'])             # encode_type['encoding'] = EUC-KR
                    line = line.split('\t', maxsplit=3)
                    wr.writerow(line)
                else:
                    line = line.decode('EUC-KR')
            except:
                print("problem is occured. id", id)
    f.close()

def main():

    global row_cr
    global row_tc
    global user_id_pw
    global implementation_criteria
    global minor_criteria
    global LukID_temp
    global SwID_length
    global SysID_length
    global cr_delivery_milestone
    global TC_Review_Status
    global read_xlsx
    global test_session_txt
    global TestResult_csv
    global write_xlsx


    implementation_criteria = 12
    row_temp_first = 2
    row_temp = 2
    minor_criteria = 'Patch#2'

    user_id_pw = [' ', ' ']
    # Get test session IDs

    user_name_txt = "user.txt"
    DocID_Info_xls = "DocID_Info.xls"
    SysID_Info_xls = "SysID_Info.xlsx"
    SwID_Info_xls = "SwID_Info.xlsx"
    SysSwTS_Info_xls = "SysSwTSID_Info.xlsx"                         ### 고정

    ########################## user.txt 파일 읽기 ##########################
    try:
        with open(user_name_txt, 'rt') as in_file:
            for line in in_file:
                user_id_pw = line                               # user_name_txt 파일에 데이터 1줄마다 읽음
                user_id_pw = user_id_pw.split(',')              # 데이터의 id와 pw ,로 구분
                user_id_pw[0] = user_id_pw[0].strip()           # user_id_pw[0]에 id 데이터 저장
                user_id_pw[1] = user_id_pw[1].strip()           # user_id_pw[1]에 pw 데이터 저장
    except:
        print("Please make a user.txt file with ID, PW. e.g. ID, PW")
        subprocess.check_output("pause", shell=True)
        return

    Result_DocID = 'DocID_Info.xls'
    Result_SysRS1 = 'SysRS1_Info.xls'
    Result_SysRS2 = 'SysRS2_Info.xls'
    Result_SysRS3 = 'SysRS3_Info.xls'
    Result_SwRS1 = 'SwRS1_Info.xls'
    Result_SwRS2 = 'SwRS2_Info.xls'
    Result_SwRS3 = 'SwRS3_Info.xls'
    Result_SysSwTS1 = 'SysTC1_Info.xls'
    Result_SysSwTS2 = 'SysTC2_Info.xls'
    Result_SysSwTS3 = 'SysTC3_Info.xls'

    if os.path.exists(Result_DocID):
        os.remove(Result_DocID)
    if os.path.exists(Result_SysRS1):
        os.remove(Result_SysRS1)
    if os.path.exists(Result_SysRS2):
        os.remove(Result_SysRS2)
    if os.path.exists(Result_SysRS3):
        os.remove(Result_SysRS3)
    if os.path.exists(Result_SwRS1):
        os.remove(Result_SwRS1)
    if os.path.exists(Result_SwRS2):
        os.remove(Result_SwRS2)
    if os.path.exists(Result_SwRS3):
        os.remove(Result_SwRS3)
    if os.path.exists(Result_SysSwTS1):
        os.remove(Result_SysSwTS1)
    if os.path.exists(Result_SysSwTS2):
        os.remove(Result_SysSwTS2)
    if os.path.exists(Result_SysSwTS3):
        os.remove(Result_SysSwTS3)

    QueryDefinition_DocID = '((field["Document ID"]=' + DocID + ')and(field["Project"]="/Schaeffler MCA LCU")and(item.live)and(item.meaningful)and("disabled not"(field["Category"]="Heading","Comment")))'
    QueryDefinition_SysRS1 = '((field["Document ID"]=' + SysRS1ID + ')and(field["Project"]="/Schaeffler MCA LCU")and(item.live)and(item.meaningful)and("disabled not"(field["Category"]="Heading","Comment")))'
    QueryDefinition_SysRS2 = '((field["Document ID"]=' + SysRS2ID + ')and(field["Project"]="/Schaeffler MCA LCU")and(item.live)and(item.meaningful)and("disabled not"(field["Category"]="Heading","Comment")))'
    QueryDefinition_SysRS3 = '((field["Document ID"]=' + SysRS3ID + ')and(field["Project"]="/Schaeffler MCA LCU")and(item.live)and(item.meaningful)and("disabled not"(field["Category"]="Heading","Comment")))'
    QueryDefinition_SwRS1 = '((field["Document ID"]=' + SwRS1ID + ')and(field["Project"]="/Schaeffler MCA LCU")and(item.live)and(item.meaningful)and("disabled not"(field["Category"]="Heading","Comment")))'
    QueryDefinition_SwRS2 = '((field["Document ID"]=' + SwRS2ID + ')and(field["Project"]="/Schaeffler MCA LCU")and(item.live)and(item.meaningful)and("disabled not"(field["Category"]="Heading","Comment")))'
    QueryDefinition_SwRS3 = '((field["Document ID"]=' + SwRS3ID + ')and(field["Project"]="/Schaeffler MCA LCU")and(item.live)and(item.meaningful)and("disabled not"(field["Category"]="Heading","Comment")))'
    QueryDefinition_SysSwTS1 = '((field["Document ID"]=' + SysTCID + ')and(field["Project"]="/Schaeffler MCA LCU")and(item.live)and(item.meaningful)and("disabled not"(field["Category"]="Heading","Comment"))and(field["MCA OEM"]="Ferrari"))'
    QueryDefinition_SysSwTS2 = '((field["Document ID"]=' + SwTCID + ')and(field["Project"]="/Schaeffler MCA LCU")and(item.live)and(item.meaningful)and("disabled not"(field["Category"]="Heading","Comment"))and(field["MCA OEM"]="Ferrari"))'      #v2xx.x는 HMC대신 Ferrari
    QueryDefinition_SysSwTS3 = '((field["Document ID"]=' + SysITSID + ')and(field["Project"]="/Schaeffler MCA LCU")and(item.live)and(item.meaningful)and("disabled not"(field["Category"]="Heading","Comment"))and(field["MCA OEM"]="Ferrari"))'

    itemExportFields_DocID = '"Document ID",ID,"A15 LuK ID",Text,"A05 Safety Integrity","A25 Status Commitment Supplier - MCA LG","A27 Delivery Date","Decomposes To","Short Description"'
    itemExportFields_SysRS = '"Document ID",ID,"ENG ID","Validated By","Satisfied By"'
    itemExportFields_SysSwTS = '"Document ID",ID,"ENG ID","Test Method"'

    export_doc_cmd_DocID = 'im exportissues --outputFile=' + Result_DocID + ' --fields=' + itemExportFields_DocID + ' --sortField=Type --queryDefinition=' + QueryDefinition_DocID + ' --noopenOutputFile'
    export_doc_cmd_SysRS1 = 'im exportissues --outputFile=' + Result_SysRS1 + ' --fields=' + itemExportFields_SysRS + ' --sortField=Type --queryDefinition=' + QueryDefinition_SysRS1 + ' --noopenOutputFile'
    export_doc_cmd_SysRS2 = 'im exportissues --outputFile=' + Result_SysRS2 + ' --fields=' + itemExportFields_SysRS + ' --sortField=Type --queryDefinition=' + QueryDefinition_SysRS2 + ' --noopenOutputFile'
    export_doc_cmd_SysRS3 = 'im exportissues --outputFile=' + Result_SysRS3 + ' --fields=' + itemExportFields_SysRS + ' --sortField=Type --queryDefinition=' + QueryDefinition_SysRS3 + ' --noopenOutputFile'
    export_doc_cmd_SwRS1 = 'im exportissues --outputFile=' + Result_SwRS1 + ' --fields=' + itemExportFields_SysRS + ' --sortField=Type --queryDefinition=' + QueryDefinition_SwRS1 + ' --noopenOutputFile'
    export_doc_cmd_SwRS2 = 'im exportissues --outputFile=' + Result_SwRS2 + ' --fields=' + itemExportFields_SysRS + ' --sortField=Type --queryDefinition=' + QueryDefinition_SwRS2 + ' --noopenOutputFile'
    export_doc_cmd_SwRS3 = 'im exportissues --outputFile=' + Result_SwRS3 + ' --fields=' + itemExportFields_SysRS + ' --sortField=Type --queryDefinition=' + QueryDefinition_SwRS3 + ' --noopenOutputFile'
    export_doc_cmd_SysSwTS1 = 'im exportissues --outputFile=' + Result_SysSwTS1 + ' --fields=' + itemExportFields_SysSwTS + ' --sortField=Type --queryDefinition=' + QueryDefinition_SysSwTS1 + ' --noopenOutputFile'
    export_doc_cmd_SysSwTS2 = 'im exportissues --outputFile=' + Result_SysSwTS2 + ' --fields=' + itemExportFields_SysSwTS + ' --sortField=Type --queryDefinition=' + QueryDefinition_SysSwTS2 + ' --noopenOutputFile'
    export_doc_cmd_SysSwTS3 = 'im exportissues --outputFile=' + Result_SysSwTS3 + ' --fields=' + itemExportFields_SysSwTS + ' --sortField=Type --queryDefinition=' + QueryDefinition_SysSwTS3 + ' --noopenOutputFile'

    start1 = time.time()

    subprocess.call(export_doc_cmd_DocID)
    subprocess.call(export_doc_cmd_SysRS1)
    subprocess.call(export_doc_cmd_SysRS2)
    subprocess.call(export_doc_cmd_SysRS3)
    subprocess.call(export_doc_cmd_SwRS1)
    subprocess.call(export_doc_cmd_SwRS2)
    subprocess.call(export_doc_cmd_SwRS3)
    subprocess.call(export_doc_cmd_SysSwTS1)
    subprocess.call(export_doc_cmd_SysSwTS2)
    subprocess.call(export_doc_cmd_SysSwTS3)

    end1 = time.time()
    print(f"기본 파일 만드는데 까지 걸리는 시간 : {end1 - start1:.5f} sec")


    ########################## SysID, SwID, SysSwTSID, TestResult_Info 파일 생성 및 통합 ##########################
    start2 = time.time()
    SysID_combined(Result_SysRS1, Result_SysRS2, Result_SysRS3)
    SwID_combined(Result_SwRS1, Result_SwRS2, Result_SwRS3)
    SysSwTSID_combined(Result_SysSwTS1, Result_SysSwTS2, Result_SysSwTS3)
    TestResult_Info()


    end2 = time.time()
    print(f"파일 통합하고 TestResult_Info 만드는데 까지 걸리는 시간 : {end2 - start2:.5f} sec")

    if os.path.exists(Result_SysRS1):
        os.remove(Result_SysRS1)
    if os.path.exists(Result_SysRS2):
        os.remove(Result_SysRS2)
    if os.path.exists(Result_SysRS3):
        os.remove(Result_SysRS3)
    if os.path.exists(Result_SwRS1):
        os.remove(Result_SwRS1)
    if os.path.exists(Result_SwRS2):
        os.remove(Result_SwRS2)
    if os.path.exists(Result_SwRS3):
        os.remove(Result_SwRS3)
    if os.path.exists(Result_SysSwTS1):
        os.remove(Result_SysSwTS1)
    if os.path.exists(Result_SysSwTS2):
        os.remove(Result_SysSwTS2)
    if os.path.exists(Result_SysSwTS3):
        os.remove(Result_SysSwTS3)

    start0 = time.time()
    ########################## testsession.txt 파일 읽기 ##########################
    try:
        with open(test_session_txt, 'rt') as in_file:
            for line in in_file:
                test_session_id = line  # test_session_txt 파일에 1줄마다 읽어드림
                test_session_id = test_session_id.split(',')  # ,로 문자열 리스트화에서 구별
            len_test_session_id = len(test_session_id)
    except:
        print("Please make a testsession.txt file with test session id. e.g. 1495334, 1495339, 1495343, 1495344")  # test_seesion_txt 파일에 아무것도 데이터가 없을때 또는 ,로 구분안갈때
        subprocess.check_output("pause", shell=True)
        return


    row_cr = row_tc = 2   # 2행부터 시작
    ########################## read.xlsx 파일 읽기 ##########################

    worksheet_number = 0                                # sheet number 지정
    worksheet_name = "Sheet" + str(worksheet_number)    # read.xlsx의 worksheet 이름

    data1 = pd.read_excel(read_xlsx, sheet_name=worksheet_name)                     # read 파일 읽기 1~6
    df_read = pd.DataFrame(data1)

    data2 = pd.read_excel(DocID_Info_xls, sheet_name=worksheet_name)
    df_DocID = pd.DataFrame(data2)

    data3 = pd.read_excel(SysID_Info_xls, sheet_name="Sheet1")
    df_SysID = pd.DataFrame(data3)

    data4 = pd.read_excel(SwID_Info_xls, sheet_name="Sheet1")
    df_SwID = pd.DataFrame(data4)

    data5 = pd.read_excel(SysSwTS_Info_xls, sheet_name="Sheet1")
    df_SysSwTS = pd.DataFrame(data5)

    data6 = pd.read_csv(TestResult_csv)
    df_TestResult = pd.DataFrame(data6)

    wb_write = Workbook()  # write.xlsx 파일 임시 파일 생성

    ws = wb_write.active
    ws.title = worksheet_name

    excel_header = ['CR Text', 'CR ID', 'ASIL Level', 'Acceptance LG against LuK requirement', 'Planned implementation milestone/date',
                    'Implementation state', 'If test: Test ID', 'Verification method',
                    'Test case review status', 'Test result' + str(test_session_id), 'Verification status', 'SysRS ID', 'SwRS ID', 'All TC Pass Status',
                    'If NOK or test not possible, add rational and write action planned']
    if (len(df_read) < 10000):
        ws.append(excel_header)

    else:                                                            # write.xlsx에 입력해야할 행이 10000개 이상일시 sheet를 한개 더 추가
        worksheet_number += 1
        worksheet_name = "Sheet" + str(worksheet_number)
        wb_write.create_sheet(worksheet_name)
        ws_write = wb_write[worksheet_name]
        ws.append(excel_header)

    #### 데이터 입력 시작 ####
    df_LuKID = df_read['A15 LuK ID']
    p = re.compile('M([\d]*)')                                      # 정규표현식을 미리 컴파일 해두는것
    for j in range(0, len(df_read)):                                # read파일에 있는 A15 LuK ID의 갯수만큼 for문
        All_TC_Pass_data = []
        DocID_Info_line = data2["A15 LuK ID"].str.contains(df_LuKID[j])
        Save_DocID_Info_line = data2[DocID_Info_line]
        if Save_DocID_Info_line.empty:
            data_row = ['WRONG CR ID', 'WRONG CR ID']
            ws.append(data_row)
            row_temp += 1
            continue
        data_row = Save_DocID_Info_line.loc[:, ['Text', "A15 LuK ID", "A05 Safety Integrity", "A25 Status Commitment Supplier - MCA LG", "A27 Delivery Date"]]
        data_row = data_row.values.tolist()[0]                         # Text, A15 LuK ID, A05 Safety Integrity, A25 Status Commitment Supplier - MCA LG, A27 Delivery Date 추출 완료//

        if str(data_row[2]) == 'nan':
            data_row[2] = 'None'

        if str(data_row[4]) == 'nan':                                          # data_row[4] 는 A27 Delivery Date에 대한 정보 M10, M11, n/a 같은 // str(data_row[4]) == 'nan' 로 바꿔야 하나?
            data_row[4] = 'in-review'
        elif data_row[4] =='all milestones':
            data_row[4] = 'all milestones'
        else:
            if data_row[4] != 'n/a':
                #if "M4_1" in data_row[4]:
                #    print("test")
                m = p.search(data_row[4])
                cr_delivery_milestone = m.group(1)
            else:
                cr_delivery_milestone = 1

        if (int(cr_delivery_milestone) != 0) & (int(cr_delivery_milestone) <= implementation_criteria):                      #implementation_criteria = 12 , cr_delivery_milesone이 0이 아니고 숫자가 12보다 작은 경우
            if minor_criteria in str(data_row[4]):                                                                           #cr_delivery = M10, M11, n/a같은 / minor_criteria = 'Patch#2'
                cr_delivery_milestone = 'not implemented'
            else:
                cr_delivery_milestone = 'implemented'                                                                        #거의 다 이쪽 실행
        else:
            cr_delivery_milestone = 'not implemented'

        data_row.append(cr_delivery_milestone)                           # cr_delivery_milestone 추출 완료//
        cr_delivery_milestone = 0

        Decom_temp = Save_DocID_Info_line.loc[:, ["Decomposes To"]]      # DecomposesTo ID 임시 저장소
        Decom_temp = Decom_temp.values.tolist()[0]
        DecomID_data_row = Decom_temp[0]

        Short_temp = Save_DocID_Info_line.loc[:, ["Short Description"]]
        Short_temp = Short_temp.values.tolist()[0]
        Short_temp = Short_temp[0]

        if str(DecomID_data_row) != 'nan':
            DecomID_data_row = DecomID_data_row.split(',')
            DecomID_length = len(DecomID_data_row)
            for i in range(0, DecomID_length):
                DecomposesToID = DecomID_data_row[i].replace("?", "").lstrip().rstrip()
                SysID_Info_line = df_SysID.loc[df_SysID['ID'] == int(DecomposesToID)]
                if SysID_Info_line.empty:
                    #ws.append(data_row)
                    #row_temp += 1
                    continue
                SysID_data_row = SysID_Info_line.loc[:, ["ENG ID", "Validated By", "Satisfied By"]]
                SysID_data_row = SysID_data_row.values.tolist()[0]

                SysRSID = SysID_data_row[0]
                SysTsID_data_row = SysID_data_row[1]
                SwID_row = SysID_data_row[2]

                if str(SysTsID_data_row) != 'nan':
                    SysTsID_data_row = SysTsID_data_row.split(',')
                    SysTsID_length = len(SysTsID_data_row)
                    for i in range(0, SysTsID_length):
                        SysTsID = SysTsID_data_row[i].replace("?", "").lstrip().rstrip()
                        SysTSID_Info_line = df_SysSwTS.loc[df_SysSwTS['ID'] == int(SysTsID)]
                        SysTSID_data_row = SysTSID_Info_line.loc[:, ["ENG ID", "Test Method"]]
                        if SysTSID_data_row.empty:
                            SysTSID_data_row = ['n/a', 'n/a']
                            TC_Review_Status = 'n/a'
                        else:
                            SysTSID_data_row = SysTSID_data_row.values.tolist()[0]
                            TC_Review_Status = 'reviewed'

                        SysTSID_temp = SysTSID_Info_line.loc[:, ["ID"]]
                        if SysTSID_temp.empty:
                            TC_Result = 'NT'
                        else:
                            SysTSID_temp = SysTSID_temp.values.tolist()[0][0]
                            TC_Result_Info_line = df_TestResult.loc[df_TestResult["TC ID"] == int(SysTSID_temp)]
                            TC_Result_data_row = TC_Result_Info_line.loc[:, ["Session ID", "Test Result"]]
                            if TC_Result_data_row.empty:
                                TC_Result = 'NT'
                            else:
                                TC_Result_data_row = TC_Result_data_row.values.tolist()[0]
                                TC_Result = TC_Result_data_row[1] + "(session : " + str(TC_Result_data_row[0]) + ")"

                        row_data = data_row + SysTSID_data_row
                        row_data.append(TC_Review_Status)
                        row_data.append(TC_Result)

                        if str('Passed') in str(TC_Result):
                            All_TC_Pass_data.append('P')
                        else:
                            All_TC_Pass_data.append('F')

                        if str('Passed') in str(TC_Result):
                            verification_status = 'finished'
                        elif str('Failed') in str(TC_Result):
                            verification_status = 'not finished'
                        else:
                            if Short_temp == 'finished':
                                verification_status = 'finished'
                            else:
                                verification_status = 'not finished'
                        row_data.append(verification_status)
                        row_data.append(SysRSID)
                        ws.append(row_data)
                        row_temp += 1

                if str(SwID_row) != 'nan':
                    SwID_row = SwID_row.split(',')
                    SwID_row = [i.replace('?', '').strip() for i in SwID_row]
                    SwID_length = len(SwID_row)
                    for i in range(0, SwID_length):
                        SwID = SwID_row[i]
                        SwID_Info_line = df_SwID.loc[df_SwID['ID'] == int(SwID)]
                        SwID_data_row = SwID_Info_line.loc[:, ["ENG ID", "Validated By"]]
                        if SwID_Info_line.empty:
                            #row_data = data_row + [' ', ' ', ' ', ' ', ' ', SysRSID]
                            #ws.append(row_data)
                            #row_temp += 1
                            continue
                        else:
                            SwID_data_row = SwID_data_row.values.tolist()[0]

                        SwRSID = SwID_data_row[0]
                        SwTsID_data_row = SwID_data_row[1]

                        if str(SwTsID_data_row) != 'nan':
                            SwTsID_ValidatedBy = SwTsID_data_row.split(",")
                            SwTsID_length = len(SwTsID_ValidatedBy)
                            for i in range(0, SwTsID_length):
                                SwTsID = SwTsID_ValidatedBy[i].replace("?", "").lstrip().rstrip()
                                SwTSID_Info_line = df_SysSwTS.loc[df_SysSwTS['ID'] == int(SwTsID)]
                                SwTSID_data_row = SwTSID_Info_line.loc[:, ["ENG ID", "Test Method"]]
                                if SwTSID_data_row.empty:
                                    SwTSID_data_row = ['n/a', 'n/a']
                                    TC_Review_Status = 'n/a'
                                else:
                                    SwTSID_data_row = SwTSID_data_row.values.tolist()[0]
                                    TC_Review_Status = 'reviewed'

                                SwTSID_temp = SwTSID_Info_line.loc[:, ["ID"]]
                                if SwTSID_temp.empty:
                                    TC_Result = 'NT'
                                else:
                                    SwTSID_temp = SwTSID_temp.values.tolist()[0][0]
                                    TC_Result_Info_line = df_TestResult.loc[df_TestResult["TC ID"] == int(SwTSID_temp)]
                                    TC_Result_data_row = TC_Result_Info_line.loc[:, ["Session ID", "Test Result"]]
                                    if TC_Result_data_row.empty:
                                        TC_Result = 'NT'
                                    else:
                                        TC_Result_data_row = TC_Result_data_row.values.tolist()[0]
                                        TC_Result = TC_Result_data_row[1] + "(session : " + str(TC_Result_data_row[0]) + ")"

                                row_data = data_row + SwTSID_data_row
                                row_data.append(TC_Review_Status)
                                row_data.append(TC_Result)

                                if str('Passed') in str(TC_Result):
                                    All_TC_Pass_data.append('P')
                                else:
                                    All_TC_Pass_data.append('F')

                                if str('Passed') in str(TC_Result):
                                    verification_status = 'finished'
                                elif str('Failed') in str(TC_Result):
                                    verification_status = 'not finished'
                                else:
                                    if Short_temp == 'finished':
                                        verification_status = 'finished'
                                    else:
                                        verification_status = 'not finished'

                                row_data.append(verification_status)
                                row_data.append(' ')
                                row_data.append(SwRSID)
                                ws.append(row_data)
                                row_temp += 1

        if 'F' in All_TC_Pass_data:
            All_TC_Pass = "All TC Not Pass" + '[' + str(df_LuKID[j]) + ']'
        else:
            All_TC_Pass = "All TC Pass" + '[' + str(df_LuKID[j]) + ']'

        if j == 0:
            row_temp = row_temp_first

        ws.cell(row_temp, 14).value = All_TC_Pass

    #### 데이터 입력 종료 ####
    wb_write.save(write_xlsx)
    wb_write.close()
    end0 = time.time()
    print(f"Write파일이 만드는데 걸리는 시간 : {end0 - start0:.5f} sec")


##################################################################
main()

end1 = time.time()
print(f"Write파일이 만드는데 까지 걸리는 총 시간 : {end1 - start1:.5f} sec")