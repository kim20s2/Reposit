"""
CopyRight ⓒ 2022, SungHwan Kim

2022. 01. 19 Ver. 0.2 completed

This library is free software;
"""
#import numpy.random.common
#import numpy.random.bounded_integers
#import numpy.random.entropy
import pandas as pd
import sys
from PyQt5 import *
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
from PyQt5.QtGui import QIcon
import openpyxl
import win32com.client
from openpyxl import Workbook
import subprocess
import os
import re
import chardet
from enum import Enum
import csv
import xlwt

test_session_id = []
len_test_session_id = 0
TestResult_csv = "TestResult_Info.csv"

implementation_criteria = 0
minor_criteria = 0
test_result = 0
LukID_temp = 0
DecomposesToID = 0
SysID_length = 0
SwID_length = 0
cr_short_description = 0
cr_delivery_milestone = 0
verification_status = 'not finished'
TC_Review_Status = 0
write_xlsx = ""
OEM_Filter = ""
OEM_Filter_Error1 = 0
OEM_Filter_Error2 = 0
OEM_Filter_Error_Signal = 0
read_xlsx = ""


class App(QMainWindow):

    def __init__(self):
        super().__init__()
        self.date = QDate.currentDate()
        self.initUI()


    def initUI(self):
        global cb1
        global cb2

        self.setWindowTitle("Trace Matrix Making Tool (Ver. 0.2)")
        self.setWindowIcon(QIcon('web.png'))
        self.setFixedSize(490, 350)

        exitAction = QAction('Exit', self)
        exitAction.setShortcut('Ctrl+Q')
        exitAction.setStatusTip('Exit application')
        exitAction.triggered.connect(qApp.quit)

        self.statusBar().showMessage(self.date.toString(Qt.DefaultLocaleLongDate))
        text_label = QLabel(self)
        text_label.move(370, 325)
        text_label.setText("Made By S.H.Kim")

        menubar = self.menuBar()
        menubar.setNativeMenuBar(False)
        filemenu = menubar.addMenu('&File')
        filemenu.addAction(exitAction)


        text_label = QLabel(self)
        text_label.move(150, 90)
        text_label.setText("OEM Filter 여부")


        cb1 = QCheckBox("HMC", self)
        cb1.move(150, 110)
        cb1.clicked.connect(self.EnableOEM_HMC_Filter)

        cb2 = QCheckBox("Ferrari", self)
        cb2.move(220, 110)
        cb2.clicked.connect(self.EnableOEM_Ferrari_Filter)


        text_label = QLabel(self)
        text_label.move(10, 20)
        text_label.setText('User ID')
        self.line_UserID = QLineEdit(self)
        self.line_UserID.move(10, 45)

        text_label = QLabel(self)
        text_label.move(110, 20)
        text_label.setText('Password')
        self.line_PW = QLineEdit(self)
        self.line_PW.move(110, 45)

        text_label = QLabel(self)
        text_label.move(10, 90)
        text_label.setText('CR DocID')
        self.line_DocID = QLineEdit(self)
        self.line_DocID.move(10, 115)

        text_label = QLabel(self)
        text_label.move(10, 140)
        text_label.setText('SysRS #1')
        self.line_SysRS1 = QLineEdit(self)
        self.line_SysRS1.move(10, 165)

        text_label = QLabel(self)
        text_label.move(110, 140)
        text_label.setText('SysRS #2')
        self.line_SysRS2 = QLineEdit(self)
        self.line_SysRS2.move(110, 165)

        text_label = QLabel(self)
        text_label.move(210, 140)
        text_label.setText('SysRS #3')
        self.line_SysRS3 = QLineEdit(self)
        self.line_SysRS3.move(210, 165)

        text_label = QLabel(self)
        text_label.move(10, 190)
        text_label.setText('SwRS #1')
        self.line_SwRS1 = QLineEdit(self)
        self.line_SwRS1.move(10, 215)

        text_label = QLabel(self)
        text_label.move(110, 190)
        text_label.setText('SwRS #2')
        self.line_SwRS2 = QLineEdit(self)
        self.line_SwRS2.move(110, 215)

        text_label = QLabel(self)
        text_label.move(210, 190)
        text_label.setText('SwRS #3')
        self.line_SwRS3 = QLineEdit(self)
        self.line_SwRS3.move(210, 215)

        text_label = QLabel(self)
        text_label.move(10, 240)
        text_label.setText('SwTS')
        self.line_SwTC = QLineEdit(self)
        self.line_SwTC.move(10, 265)

        text_label = QLabel(self)
        text_label.move(110, 240)
        text_label.setText('SysITS')
        self.line_SysITS = QLineEdit(self)
        self.line_SysITS.move(110, 265)

        text_label = QLabel(self)
        text_label.move(210, 240)
        text_label.setText('SysTS')
        self.line_SysTC = QLineEdit(self)
        self.line_SysTC.move(210, 265)

        text_label = QLabel(self)
        text_label.move(340, 190)
        text_label.setText('출력 파일명 입력')
        self.line_ResultName = QLineEdit(self)
        self.line_ResultName.move(340, 215)

        btn_ID = QPushButton('ID 저장', self)
        btn_ID.clicked.connect(self.btn_ID)
        btn_ID.move(230, 40)
        btn_ID.resize(btn_ID.sizeHint())

        btn_ID_upload = QPushButton('ID 불러오기', self)
        btn_ID_upload.clicked.connect(self.btn_ID_upload)
        btn_ID_upload.move(230, 65)
        btn_ID_upload.resize(btn_ID_upload.sizeHint())

        text_label = QLabel(self)
        text_label.move(330, 35)
        text_label.setText('①')
        btn1 = QPushButton('Integrity Server 로그인', self)
        btn1.clicked.connect(self.btn1_clicked)
        btn1.move(350, 40)
        btn1.resize(btn1.sizeHint())

        text_label = QLabel(self)
        text_label.move(330, 65)
        text_label.setText('②')
        btn2 = QPushButton('Base File 생성', self)
        btn2.clicked.connect(self.btn2_clicked)
        btn2.move(350, 70)
        btn2.resize(btn2.sizeHint())

        text_label = QLabel(self)
        text_label.move(330, 95)
        text_label.setText('③')
        btn3 = QPushButton('Testsession.txt 읽기', self)
        btn3.clicked.connect(self.btn3_clicked)
        btn3.move(350, 100)
        btn3.resize(btn3.sizeHint())

        text_label = QLabel(self)
        text_label.move(330, 125)
        text_label.setText('④')
        btn4 = QPushButton('Read File 읽기', self)
        btn4.clicked.connect(self.btn4_clicked)
        btn4.move(350, 130)
        btn4.resize(btn4.sizeHint())

        text_label = QLabel(self)
        text_label.move(330, 155)
        text_label.setText('⑤')
        btn5 = QPushButton('TestResult File 생성', self)
        btn5.clicked.connect(self.btn5_clicked)
        btn5.move(350, 160)
        btn5.resize(btn5.sizeHint())

        text_label = QLabel(self)
        text_label.move(330, 260)
        text_label.setText('⑥')
        btn6 = QPushButton('Trace Matrix 생성', self)
        btn6.clicked.connect(self.btn6_clicked)
        btn6.move(350, 265)
        btn6.resize(btn6.sizeHint())


    def EnableOEM_HMC_Filter(self):
        global OEM_Filter
        global OEM_Filter_Error1
        global OEM_Filter_Error2

        if cb1.isChecked():
            OEM_Filter = """field["MCA OEM"]="HMC")"""
            OEM_Filter_Error1 = 1
        else:
            OEM_Filter = ""
            OEM_Filter_Error1 = 0

        if (OEM_Filter_Error1 * OEM_Filter_Error2) == 1:
            QMessageBox.about(self, "Warning", "HMC 또는 Ferrari중 하나만 선택하세요!")
            cb1.toggle()
            cb2.toggle()
            OEM_Filter = ""
            OEM_Filter_Error1 = 0
            OEM_Filter_Error2 = 0

    def EnableOEM_Ferrari_Filter(self):
        global OEM_Filter
        global OEM_Filter_Error1
        global OEM_Filter_Error2

        if cb2.isChecked():
            OEM_Filter = """field["MCA OEM"]="Ferrari")"""
            OEM_Filter_Error2 = 1
        else:
            OEM_Filter = ""
            OEM_Filter_Error2 = 0

        if (OEM_Filter_Error1 * OEM_Filter_Error2) == 1:
            QMessageBox.about(self, "Warning", "HMC 또는 Ferrari중 하나만 선택하세요!")
            cb1.toggle()
            cb2.toggle()
            OEM_Filter = ""
            OEM_Filter_Error1 = 0
            OEM_Filter_Error2 = 0

    def btn_ID(self):
        f = open('Saved ID.txt', 'w')
        f.write(str(self.line_DocID.text()) + '\n')
        f.write(str(self.line_SysRS1.text()) + '\n')
        f.write(str(self.line_SysRS2.text()) + '\n')
        f.write(str(self.line_SysRS3.text()) + '\n')
        f.write(str(self.line_SwRS1.text()) + '\n')
        f.write(str(self.line_SwRS2.text()) + '\n')
        f.write(str(self.line_SwRS3.text()) + '\n')
        f.write(str(self.line_SwTC.text()) + '\n')
        f.write(str(self.line_SysITS.text()) + '\n')
        f.write(str(self.line_SysTC.text()) + '\n')
        f.close
        QMessageBox.about(self, "Warning", "ID 저장 완료")

    def btn_ID_upload(self):
        fileOpen = QFileDialog.getOpenFileName(self, 'Open file', './')
        if fileOpen[0]:
            f = open(fileOpen[0], 'r')
            with f:
                line = f.readlines()
        else:
            return

        for i in range(len(line)):
            line[i] = line[i].replace('\n', "")

        self.line_DocID.setText(line[0])
        self.line_SysRS1.setText(line[1])
        self.line_SysRS2.setText(line[2])
        self.line_SysRS3.setText(line[3])
        self.line_SwRS1.setText(line[4])
        self.line_SwRS2.setText(line[5])
        self.line_SwRS3.setText(line[6])
        self.line_SwTC.setText(line[7])
        self.line_SysITS.setText(line[8])
        self.line_SysTC.setText(line[9])


    def btn1_clicked(self):
        if self.line_UserID.text() != "":
            user_id_pw0 = str(self.line_UserID.text())
            temp1 = 1
        else:
            QMessageBox.about(self, "Warning", "User ID가 입력되지 않았습니다.")
            temp1 = 0
        if self.line_PW.text() != "":
            user_id_pw1 = str(self.line_PW.text())
            temp2 = 1
        else:
            QMessageBox.about(self, "Warning", "Password가 입력되지 않았습니다.")
            temp2 = 0

        if temp1 * temp2 == 1:
            connect_command = 'si connect --user=' + user_id_pw0 + ' --password=' + user_id_pw1
            subprocess.call(connect_command, shell=True)

        #### actions ####
    def btn2_clicked(self):
        global OEM_Filter

        if self.line_DocID.text() != "":
            DocID = str(self.line_DocID.text())
        if self.line_SysRS1.text() != "":
            SysRS1ID = str(self.line_SysRS1.text())
        if self.line_SysRS2.text() != "":
            SysRS2ID = str(self.line_SysRS2.text())
        if self.line_SysRS3.text() != "":
            SysRS3ID = str(self.line_SysRS3.text())
        if self.line_SwRS1.text() != "":
            SwRS1ID = str(self.line_SwRS1.text())
        if self.line_SwRS2.text() != "":
            SwRS2ID = str(self.line_SwRS2.text())
        if self.line_SwRS3.text() != "":
            SwRS3ID = str(self.line_SwRS3.text())
        if self.line_SysTC.text() != "":
            SysTCID = str(self.line_SysTC.text())
        if self.line_SwTC.text() != "":
            SwTCID = str(self.line_SwTC.text())
        if self.line_SysITS.text() != "":
            SysITSID = str(self.line_SysITS.text())

        Result_DocID = 'DocID_Info.xls'
        Result_SysRS1 = 'SysRS1_Info.xls'
        Result_SysRS2 = 'SysRS2_Info.xls'
        Result_SysRS3 = 'SysRS3_Info.xls'
        Result_SwRS1 = 'SwRS1_Info.xls'
        Result_SwRS2 = 'SwRS2_Info.xls'
        Result_SwRS3 = 'SwRS3_Info.xls'
        Result_SysSwTS1 = 'SysTC1_Info.xls'
        Result_SysSwTS2 = 'SysTC2_Info.xls'
        Result_SysSwTS3 = 'SysTC3_Info.xls'

        ########### 파일 3개 통합 ###########
        def SysID_combined(input1, input2, input3):
            #엑셀 파일 이름
            excel_names = [input1, input2, input3]
            excel_names = [v for v in excel_names if v]
            excels = [pd.ExcelFile(name) for name in excel_names]
            frames = [x.parse(x.sheet_names[0], header=None,index_col=None) for x in excels]
            frames[1:] = [df[1:] for df in frames[1:]]
            combined = pd.concat(frames)
            temp_list = (combined.iloc[0]).values.tolist()
            if 'ID' not in temp_list:
                header = pd.DataFrame([["Document ID",'ID',"A15 LuK ID",'Text',"A05 Safety Integrity","A25 Status Commitment Supplier - MCA LG","A27 Delivery Date","Decomposes To","Short Description"]])
                combined = pd.concat([header, combined], ignore_index=True)
            #파일저장
            combined.to_excel("SysID_Info.xlsx", header=False, index=False)

        def SwID_combined(input1, input2, input3):
            #엑셀 파일 이름
            excel_names = [input1, input2, input3]
            excel_names = [v for v in excel_names if v]
            excels = [pd.ExcelFile(name) for name in excel_names]
            frames = [x.parse(x.sheet_names[0], header=None,index_col=None) for x in excels]
            frames[1:] = [df[1:] for df in frames[1:]]
            combined = pd.concat(frames)
            temp_list = (combined.iloc[0]).values.tolist()
            if 'ID' not in temp_list:
                header = pd.DataFrame([["Document ID",'ID',"ENG ID","Validated By","Satisfied By"]])
                combined = pd.concat([header, combined], ignore_index=True)
            #파일저장
            combined.to_excel("SwID_Info.xlsx", header=False, index=False)

        def SysSwTSID_combined(input1, input2, input3):
            #엑셀 파일 이름
            excel_names = [input1, input2, input3]
            excel_names = [v for v in excel_names if v]
            excels = [pd.ExcelFile(name) for name in excel_names]
            frames = [x.parse(x.sheet_names[0], header=None,index_col=None) for x in excels]
            frames[1:] = [df[1:] for df in frames[1:]]
            combined = pd.concat(frames)
            temp_list = (combined.iloc[0]).values.tolist()
            if "ID" not in temp_list:
                header = pd.DataFrame([["Document ID",'ID',"ENG ID","Test Method"]])
                combined = pd.concat([header, combined], ignore_index=True)
            #파일저장
            combined.to_excel("SysSwTSID_Info.xlsx", header=False, index=False)

        ########## 기존 파일 삭제 ##########
        if os.path.exists(Result_DocID):
            os.remove(Result_DocID)
        if os.path.exists(Result_SysRS1):
            os.remove(Result_SysRS1)
        if os.path.exists(Result_SysRS2):
            os.remove(Result_SysRS2)
        if os.path.exists(Result_SysRS3):
            os.remove(Result_SysRS3)
        if os.path.exists(Result_SwRS1):
            os.remove(Result_SwRS1)
        if os.path.exists(Result_SwRS2):
            os.remove(Result_SwRS2)
        if os.path.exists(Result_SwRS3):
            os.remove(Result_SwRS3)
        if os.path.exists(Result_SysSwTS1):
            os.remove(Result_SysSwTS1)
        if os.path.exists(Result_SysSwTS2):
            os.remove(Result_SysSwTS2)
        if os.path.exists(Result_SysSwTS3):
            os.remove(Result_SysSwTS3)

        itemExportFields_DocID = '"Document ID",ID,"A15 LuK ID",Text,"A05 Safety Integrity","A25 Status Commitment Supplier - MCA LG","A27 Delivery Date","Decomposes To","Short Description"'
        itemExportFields_SysRS = '"Document ID",ID,"ENG ID","Validated By","Satisfied By"'
        itemExportFields_SysSwTS = '"Document ID",ID,"ENG ID","Test Method"'

        ############# 여기서 파일 생성 실행 ##############
        if self.line_DocID.text() != "":
            QueryDefinition_DocID = '((field["Document ID"]=' + DocID + ')and(field["Project"]="/Schaeffler MCA LCU")and(item.live)and(item.meaningful)and("disabled not"(field["Category"]="Heading","Comment")))'
            export_doc_cmd_DocID = 'im exportissues --outputFile=' + Result_DocID + ' --fields=' + itemExportFields_DocID + ' --sortField=Type --queryDefinition=' + QueryDefinition_DocID + ' --noopenOutputFile'
            subprocess.call(export_doc_cmd_DocID)
        else:
            Result_DocID = ""
            QMessageBox.about(self, "Warning", "CR DocID가 입력되지 않았습니다.")
            return
        if self.line_SysRS1.text() != "":
            QueryDefinition_SysRS1 = '((field["Document ID"]=' + SysRS1ID + ')and(field["Project"]="/Schaeffler MCA LCU")and(item.live)and(item.meaningful)and("disabled not"(field["Category"]="Heading","Comment")))'
            export_doc_cmd_SysRS1 = 'im exportissues --outputFile=' + Result_SysRS1 + ' --fields=' + itemExportFields_SysRS + ' --sortField=Type --queryDefinition=' + QueryDefinition_SysRS1 + ' --noopenOutputFile'
            subprocess.call(export_doc_cmd_SysRS1)
        else:
            Result_SysRS1 = ""
        if self.line_SysRS2.text() != "":
            QueryDefinition_SysRS2 = '((field["Document ID"]=' + SysRS2ID + ')and(field["Project"]="/Schaeffler MCA LCU")and(item.live)and(item.meaningful)and("disabled not"(field["Category"]="Heading","Comment")))'
            export_doc_cmd_SysRS2 = 'im exportissues --outputFile=' + Result_SysRS2 + ' --fields=' + itemExportFields_SysRS + ' --sortField=Type --queryDefinition=' + QueryDefinition_SysRS2 + ' --noopenOutputFile'
            subprocess.call(export_doc_cmd_SysRS2)
        else:
            Result_SysRS2 = ""
        if self.line_SysRS3.text() != "":
            QueryDefinition_SysRS3 = '((field["Document ID"]=' + SysRS3ID + ')and(field["Project"]="/Schaeffler MCA LCU")and(item.live)and(item.meaningful)and("disabled not"(field["Category"]="Heading","Comment")))'
            export_doc_cmd_SysRS3 = 'im exportissues --outputFile=' + Result_SysRS3 + ' --fields=' + itemExportFields_SysRS + ' --sortField=Type --queryDefinition=' + QueryDefinition_SysRS3 + ' --noopenOutputFile'
            subprocess.call(export_doc_cmd_SysRS3)
        else:
            Result_SysRS3 = ""
        if self.line_SwRS1.text() != "":
            QueryDefinition_SwRS1 = '((field["Document ID"]=' + SwRS1ID + ')and(field["Project"]="/Schaeffler MCA LCU")and(item.live)and(item.meaningful)and("disabled not"(field["Category"]="Heading","Comment")))'
            export_doc_cmd_SwRS1 = 'im exportissues --outputFile=' + Result_SwRS1 + ' --fields=' + itemExportFields_SysRS + ' --sortField=Type --queryDefinition=' + QueryDefinition_SwRS1 + ' --noopenOutputFile'
            subprocess.call(export_doc_cmd_SwRS1)
        else:
            Result_SwRS1 = ""
        if self.line_SwRS2.text() != "":
            QueryDefinition_SwRS2 = '((field["Document ID"]=' + SwRS2ID + ')and(field["Project"]="/Schaeffler MCA LCU")and(item.live)and(item.meaningful)and("disabled not"(field["Category"]="Heading","Comment")))'
            export_doc_cmd_SwRS2 = 'im exportissues --outputFile=' + Result_SwRS2 + ' --fields=' + itemExportFields_SysRS + ' --sortField=Type --queryDefinition=' + QueryDefinition_SwRS2 + ' --noopenOutputFile'
            subprocess.call(export_doc_cmd_SwRS2)
        else:
            Result_SwRS2 = ""
        if self.line_SwRS3.text() != "":
            QueryDefinition_SwRS3 = '((field["Document ID"]=' + SwRS3ID + ')and(field["Project"]="/Schaeffler MCA LCU")and(item.live)and(item.meaningful)and("disabled not"(field["Category"]="Heading","Comment")))'
            export_doc_cmd_SwRS3 = 'im exportissues --outputFile=' + Result_SwRS3 + ' --fields=' + itemExportFields_SysRS + ' --sortField=Type --queryDefinition=' + QueryDefinition_SwRS3 + ' --noopenOutputFile'
            subprocess.call(export_doc_cmd_SwRS3)
        else:
            Result_SwRS3 = ""
        if self.line_SysTC.text() != "":
            QueryDefinition_SysSwTS1 = '((field["Document ID"]=' + SysTCID + ')and(field["Project"]="/Schaeffler MCA LCU")and(item.live)and(item.meaningful)and("disabled not"(field["Category"]="Heading","Comment"))' + OEM_Filter + ')'
            export_doc_cmd_SysSwTS1 = 'im exportissues --outputFile=' + Result_SysSwTS1 + ' --fields=' + itemExportFields_SysSwTS + ' --sortField=Type --queryDefinition=' + QueryDefinition_SysSwTS1 + ' --noopenOutputFile'
            subprocess.call(export_doc_cmd_SysSwTS1)
        else:
            Result_SysSwTS1 = ""
        if self.line_SwTC.text() != "":
            QueryDefinition_SysSwTS2 = '((field["Document ID"]=' + SwTCID + ')and(field["Project"]="/Schaeffler MCA LCU")and(item.live)and(item.meaningful)and("disabled not"(field["Category"]="Heading","Comment"))' + OEM_Filter + ')'      #v2xx.x는 HMC대신 Ferrari
            export_doc_cmd_SysSwTS2 = 'im exportissues --outputFile=' + Result_SysSwTS2 + ' --fields=' + itemExportFields_SysSwTS + ' --sortField=Type --queryDefinition=' + QueryDefinition_SysSwTS2 + ' --noopenOutputFile'
            subprocess.call(export_doc_cmd_SysSwTS2)
        else:
            Result_SysSwTS2 = ""
        if self.line_SysITS.text() != "":
            QueryDefinition_SysSwTS3 = '((field["Document ID"]=' + SysITSID + ')and(field["Project"]="/Schaeffler MCA LCU")and(item.live)and(item.meaningful)and("disabled not"(field["Category"]="Heading","Comment"))' + OEM_Filter + ')'
            export_doc_cmd_SysSwTS3 = 'im exportissues --outputFile=' + Result_SysSwTS3 + ' --fields=' + itemExportFields_SysSwTS + ' --sortField=Type --queryDefinition=' + QueryDefinition_SysSwTS3 + ' --noopenOutputFile'
            subprocess.call(export_doc_cmd_SysSwTS3)
        else:
            Result_SysSwTS3 = ""

        ############## SysID, SwID, SysSwTSID 파일 통합 ##############
        SysID_combined(Result_SysRS1, Result_SysRS2, Result_SysRS3)
        SwID_combined(Result_SwRS1, Result_SwRS2, Result_SwRS3)
        SysSwTSID_combined(Result_SysSwTS1, Result_SysSwTS2, Result_SysSwTS3)

        if os.path.exists(Result_SysRS1):
            os.remove(Result_SysRS1)
        if os.path.exists(Result_SysRS2):
            os.remove(Result_SysRS2)
        if os.path.exists(Result_SysRS3):
            os.remove(Result_SysRS3)
        if os.path.exists(Result_SwRS1):
            os.remove(Result_SwRS1)
        if os.path.exists(Result_SwRS2):
            os.remove(Result_SwRS2)
        if os.path.exists(Result_SwRS3):
            os.remove(Result_SwRS3)
        if os.path.exists(Result_SysSwTS1):
            os.remove(Result_SysSwTS1)
        if os.path.exists(Result_SysSwTS2):
            os.remove(Result_SysSwTS2)
        if os.path.exists(Result_SysSwTS3):
            os.remove(Result_SysSwTS3)

        QMessageBox.about(self, "Base File 알림", "Base File 생성완료")


    def btn3_clicked(self):
        global test_session_id
        global len_test_session_id
        fileOpen = QFileDialog.getOpenFileName(self, 'Open file', './')
        if fileOpen[0]:
            f = open(fileOpen[0], 'r')
            with f:
                data = f.read()
                try:
                    test_session_id = data  # test_session_txt 파일에 1줄마다 읽어드림
                    test_session_id = test_session_id.split(',')  # ,로 문자열 리스트화에서 구별
                    len_test_session_id = len(test_session_id)
                except:
                    print("Please make a testsession.txt file with test session id. e.g. 1495334, 1495339, 1495343, 1495344")  # test_seesion_txt 파일에 아무것도 데이터가 없을때 또는 ,로 구분안갈때
                    subprocess.check_output("pause", shell=True)
                print(data)
                print(len_test_session_id)
                QMessageBox.about(self, "TestSession 알림", "TestSession 읽기완료")
        else:
            return

    def btn4_clicked(self):
        global read_xlsx
        read_file = QFileDialog.getOpenFileName(self, 'Open file', './')
        if read_file[0]:
            read_xlsx = read_file[0]
            print(read_xlsx)
            QMessageBox.about(self, "Read file 읽기알림", "Read file 읽기완료")
        else:
            return



    def btn5_clicked(self):
        global test_session_id
        global len_test_session_id
        global TestResult_csv

        f = open(TestResult_csv, 'w', encoding='EUC-KR', newline='')
        wr = csv.writer(f)
        wr.writerow(["Session ID", "TC ID", "Test Result"])

        itemExportFields_TestResult = "sessionID,caseID,verdict"

        for j in range(0, len_test_session_id):
            result_test_cmd = 'tm results' + ' --sessionID=' + test_session_id[j].strip() +' --fields=' + itemExportFields_TestResult

            result = subprocess.check_output(result_test_cmd)
            result = result.splitlines()

            for line in result:  # Store each line in a string variable "line"
                # parse id decomposed to, satisfied by, validated by
                try:
                    encode_type = chardet.detect(line)
                    if encode_type['encoding'] is not None:
                        line = line.decode(encode_type['encoding'])             # encode_type['encoding'] = EUC-KR
                        line = line.split('\t', maxsplit=3)
                        wr.writerow(line)
                    else:
                        line = line.decode('EUC-KR')
                except:
                    print("problem is occured. id", id)
        f.close()
        QMessageBox.about(self, "TestResult 알림", "TestResult 생성 완료")


    def btn6_clicked(self):
        global row_cr
        global row_tc
        global user_id_pw
        global implementation_criteria
        global minor_criteria
        global LukID_temp
        global SwID_length
        global SysID_length
        global cr_delivery_milestone
        global TC_Review_Status
        global TestResult_csv
        global write_xlsx
        global read_xlsx

        if self.line_ResultName.text() != "":
            write_xlsx = str(self.line_ResultName.text()) + ".xlsx"
        else:
            QMessageBox.about(self, "출력 파일명 입력 알림", "출력 파일명을 입력하세요.")

        implementation_criteria = 12
        row_temp_first = 2
        row_temp = 2
        minor_criteria = 'Patch#2'

        DocID_Info_xls = "DocID_Info.xls"
        SysID_Info_xls = "SysID_Info.xlsx"
        SwID_Info_xls = "SwID_Info.xlsx"
        SysSwTS_Info_xls = "SysSwTSID_Info.xlsx"                         ### 고정

        row_cr = row_tc = 2   # 2행부터 시작
        ########################## read.xlsx 파일 읽기 ##########################

        worksheet_number = 0                                # sheet number 지정
        worksheet_name = "Sheet" + str(worksheet_number)    # read.xls의 worksheet 이름

        data1 = pd.read_excel(read_xlsx, sheet_name=worksheet_name)                     # read 파일 읽기 1~6
        df_read = pd.DataFrame(data1)

        data2 = pd.read_excel(DocID_Info_xls, sheet_name=worksheet_name)
        df_DocID = pd.DataFrame(data2)

        data3 = pd.read_excel(SysID_Info_xls, sheet_name="Sheet1")
        df_SysID = pd.DataFrame(data3)

        data4 = pd.read_excel(SwID_Info_xls, sheet_name="Sheet1")
        df_SwID = pd.DataFrame(data4)

        data5 = pd.read_excel(SysSwTS_Info_xls, sheet_name="Sheet1")
        df_SysSwTS = pd.DataFrame(data5)

        data6 = pd.read_csv(TestResult_csv)
        df_TestResult = pd.DataFrame(data6)

        wb_write = Workbook()  # write.xlsx 파일 임시 파일 생성

        ws = wb_write.active
        ws.title = worksheet_name

        excel_header = ['CR Text', 'CR ID', 'ASIL Level', 'Acceptance LG against LuK requirement', 'Planned implementation milestone/date',
                        'Implementation state', 'If test: Test ID', 'Verification method',
                        'Test case review status', 'Test result' + str(test_session_id), 'Verification status', 'SysRS ID', 'SwRS ID', 'All TC Pass Status',
                        'If NOK or test not possible, add rational and write action planned']
        if (len(df_read) < 10000):
            ws.append(excel_header)

        else:                                                            # write.xlsx에 입력해야할 행이 10000개 이상일시 sheet를 한개 더 추가
            worksheet_number += 1
            worksheet_name = "Sheet" + str(worksheet_number)
            wb_write.create_sheet(worksheet_name)
            ws_write = wb_write[worksheet_name]
            ws.append(excel_header)

        #### 데이터 입력 시작 ####
        df_LuKID = df_read['A15 LuK ID']
        p = re.compile('M([\d]*)')                                      # 정규표현식을 미리 컴파일 해두는것
        for j in range(0, len(df_read)):                                # read파일에 있는 A15 LuK ID의 갯수만큼 for문
            All_TC_Pass_data = []
            DocID_Info_line = data2["A15 LuK ID"].str.contains(df_LuKID[j])
            Save_DocID_Info_line = data2[DocID_Info_line]
            if Save_DocID_Info_line.empty:
                data_row = ['WRONG CR ID', 'WRONG CR ID']
                ws.append(data_row)
                row_temp += 1
                continue
            data_row = Save_DocID_Info_line.loc[:, ['Text', "A15 LuK ID", "A05 Safety Integrity", "A25 Status Commitment Supplier - MCA LG", "A27 Delivery Date"]]
            data_row = data_row.values.tolist()[0]                         # Text, A15 LuK ID, A05 Safety Integrity, A25 Status Commitment Supplier - MCA LG, A27 Delivery Date 추출 완료//

            if str(data_row[2]) == 'nan':
                data_row[2] = 'None'

            if str(data_row[4]) == 'nan':                                          # data_row[4] 는 A27 Delivery Date에 대한 정보 M10, M11, n/a 같은 // str(data_row[4]) == 'nan' 로 바꿔야 하나?
                data_row[4] = 'in-review'
            elif data_row[4] =='all milestones':
                data_row[4] = 'all milestones'
            else:
                if data_row[4] != 'n/a':
                    #if "M4_1" in data_row[4]:
                    #    print("test")
                    m = p.search(data_row[4])
                    cr_delivery_milestone = m.group(1)
                else:
                    cr_delivery_milestone = 1

            if (int(cr_delivery_milestone) != 0) & (int(cr_delivery_milestone) <= implementation_criteria):                      #implementation_criteria = 12 , cr_delivery_milesone이 0이 아니고 숫자가 12보다 작은 경우
                if minor_criteria in str(data_row[4]):                                                                           #cr_delivery = M10, M11, n/a같은 / minor_criteria = 'Patch#2'
                    cr_delivery_milestone = 'not implemented'
                else:
                    cr_delivery_milestone = 'implemented'                                                                        #거의 다 이쪽 실행
            else:
                cr_delivery_milestone = 'not implemented'

            data_row.append(cr_delivery_milestone)                           # cr_delivery_milestone 추출 완료//
            cr_delivery_milestone = 0

            Decom_temp = Save_DocID_Info_line.loc[:, ["Decomposes To"]]      # DecomposesTo ID 임시 저장소
            Decom_temp = Decom_temp.values.tolist()[0]
            DecomID_data_row = Decom_temp[0]

            Short_temp = Save_DocID_Info_line.loc[:, ["Short Description"]]
            Short_temp = Short_temp.values.tolist()[0]
            Short_temp = Short_temp[0]

            if str(DecomID_data_row) != 'nan':
                DecomID_data_row = DecomID_data_row.split(',')
                DecomID_length = len(DecomID_data_row)
                for i in range(0, DecomID_length):
                    DecomposesToID = DecomID_data_row[i].replace("?", "").lstrip().rstrip()
                    SysID_Info_line = df_SysID.loc[df_SysID['ID'] == int(DecomposesToID)]
                    if SysID_Info_line.empty:
                        #ws.append(data_row)
                        #row_temp += 1
                        continue
                    SysID_data_row = SysID_Info_line.loc[:, ["ENG ID", "Validated By", "Satisfied By"]]
                    SysID_data_row = SysID_data_row.values.tolist()[0]

                    SysRSID = SysID_data_row[0]
                    SysTsID_data_row = SysID_data_row[1]
                    SwID_row = SysID_data_row[2]

                    if str(SysTsID_data_row) != 'nan':
                        SysTsID_data_row = SysTsID_data_row.split(',')
                        SysTsID_length = len(SysTsID_data_row)
                        for i in range(0, SysTsID_length):
                            SysTsID = SysTsID_data_row[i].replace("?", "").lstrip().rstrip()
                            SysTSID_Info_line = df_SysSwTS.loc[df_SysSwTS['ID'] == int(SysTsID)]
                            SysTSID_data_row = SysTSID_Info_line.loc[:, ["ENG ID", "Test Method"]]
                            if SysTSID_data_row.empty:
                                SysTSID_data_row = ['n/a', 'n/a']
                                TC_Review_Status = 'n/a'
                            else:
                                SysTSID_data_row = SysTSID_data_row.values.tolist()[0]
                                TC_Review_Status = 'reviewed'

                            SysTSID_temp = SysTSID_Info_line.loc[:, ["ID"]]
                            if SysTSID_temp.empty:
                                TC_Result = 'NT'
                            else:
                                SysTSID_temp = SysTSID_temp.values.tolist()[0][0]
                                TC_Result_Info_line = df_TestResult.loc[df_TestResult["TC ID"] == int(SysTSID_temp)]
                                TC_Result_data_row = TC_Result_Info_line.loc[:, ["Session ID", "Test Result"]]
                                if TC_Result_data_row.empty:
                                    TC_Result = 'NT'
                                else:
                                    TC_Result_data_row = TC_Result_data_row.values.tolist()[0]
                                    TC_Result = TC_Result_data_row[1] + "(session : " + str(TC_Result_data_row[0]) + ")"

                            row_data = data_row + SysTSID_data_row
                            row_data.append(TC_Review_Status)
                            row_data.append(TC_Result)

                            if str('Passed') in str(TC_Result):
                                All_TC_Pass_data.append('P')
                            else:
                                All_TC_Pass_data.append('F')

                            if str('Passed') in str(TC_Result):
                                verification_status = 'finished'
                            elif str('Failed') in str(TC_Result):
                                verification_status = 'not finished'
                            else:
                                if Short_temp == 'finished':
                                    verification_status = 'finished'
                                else:
                                    verification_status = 'not finished'
                            row_data.append(verification_status)
                            row_data.append(SysRSID)
                            ws.append(row_data)
                            row_temp += 1

                    if str(SwID_row) != 'nan':
                        SwID_row = SwID_row.split(',')
                        SwID_row = [i.replace('?', '').strip() for i in SwID_row]
                        SwID_length = len(SwID_row)
                        for i in range(0, SwID_length):
                            SwID = SwID_row[i]
                            SwID_Info_line = df_SwID.loc[df_SwID['ID'] == int(SwID)]
                            SwID_data_row = SwID_Info_line.loc[:, ["ENG ID", "Validated By"]]
                            if SwID_Info_line.empty:
                                #row_data = data_row + [' ', ' ', ' ', ' ', ' ', SysRSID]
                                #ws.append(row_data)
                                #row_temp += 1
                                continue
                            else:
                                SwID_data_row = SwID_data_row.values.tolist()[0]

                            SwRSID = SwID_data_row[0]
                            SwTsID_data_row = SwID_data_row[1]

                            if str(SwTsID_data_row) != 'nan':
                                SwTsID_ValidatedBy = SwTsID_data_row.split(",")
                                SwTsID_length = len(SwTsID_ValidatedBy)
                                for i in range(0, SwTsID_length):
                                    SwTsID = SwTsID_ValidatedBy[i].replace("?", "").lstrip().rstrip()
                                    SwTSID_Info_line = df_SysSwTS.loc[df_SysSwTS['ID'] == int(SwTsID)]
                                    SwTSID_data_row = SwTSID_Info_line.loc[:, ["ENG ID", "Test Method"]]
                                    if SwTSID_data_row.empty:
                                        SwTSID_data_row = ['n/a', 'n/a']
                                        TC_Review_Status = 'n/a'
                                    else:
                                        SwTSID_data_row = SwTSID_data_row.values.tolist()[0]
                                        TC_Review_Status = 'reviewed'

                                    SwTSID_temp = SwTSID_Info_line.loc[:, ["ID"]]
                                    if SwTSID_temp.empty:
                                        TC_Result = 'NT'
                                    else:
                                        SwTSID_temp = SwTSID_temp.values.tolist()[0][0]
                                        TC_Result_Info_line = df_TestResult.loc[df_TestResult["TC ID"] == int(SwTSID_temp)]
                                        TC_Result_data_row = TC_Result_Info_line.loc[:, ["Session ID", "Test Result"]]
                                        if TC_Result_data_row.empty:
                                            TC_Result = 'NT'
                                        else:
                                            TC_Result_data_row = TC_Result_data_row.values.tolist()[0]
                                            TC_Result = TC_Result_data_row[1] + "(session : " + str(TC_Result_data_row[0]) + ")"

                                    row_data = data_row + SwTSID_data_row
                                    row_data.append(TC_Review_Status)
                                    row_data.append(TC_Result)

                                    if str('Passed') in str(TC_Result):
                                        All_TC_Pass_data.append('P')
                                    else:
                                        All_TC_Pass_data.append('F')

                                    if str('Passed') in str(TC_Result):
                                        verification_status = 'finished'
                                    elif str('Failed') in str(TC_Result):
                                        verification_status = 'not finished'
                                    else:
                                        if Short_temp == 'finished':
                                            verification_status = 'finished'
                                        else:
                                            verification_status = 'not finished'

                                    row_data.append(verification_status)
                                    row_data.append(' ')
                                    row_data.append(SwRSID)
                                    ws.append(row_data)
                                    row_temp += 1

            if 'F' in All_TC_Pass_data:
                All_TC_Pass = "All TC Not Pass" + '[' + str(df_LuKID[j]) + ']'
            else:
                All_TC_Pass = "All TC Pass" + '[' + str(df_LuKID[j]) + ']'

            if j == 0:
                row_temp = row_temp_first

            ws.cell(row_temp, 14).value = All_TC_Pass

        #### 데이터 입력 종료 ####
        wb_write.save(write_xlsx)
        wb_write.close()
        QMessageBox.about(self, "Trace Matrix 알림", "Trace Matrix 생성완료")



if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = App()
    ex.show()
    sys.exit(app.exec_())
